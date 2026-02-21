from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session, make_response
import csv
import datetime
import os
import json
import re
import pdfplumber
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from functools import wraps
import secrets
from datetime import datetime, timedelta
import pandas as pd
import hashlib
import logging
from logging.handlers import RotatingFileHandler
import chardet
import zipfile
import time
import uuid
import threading
import queue
from collections import Counter, defaultdict
import psutil
import traceback

app = Flask(__name__)
app.secret_key = secrets.token_hex(16)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DATA_FOLDER'] = 'data'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.permanent_session_lifetime = timedelta(days=1)

# Настройка логирования
handler = RotatingFileHandler('app.log', maxBytes=10000, backupCount=3)
handler.setFormatter(logging.Formatter(
    '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
))
handler.setLevel(logging.INFO)
app.logger.addHandler(handler)

# Создаем необходимые папки
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['DATA_FOLDER'], exist_ok=True)

# Пути к файлам данных
PRODUCTS_FILE = os.path.join(app.config['DATA_FOLDER'], 'products.csv')
ORDERS_FILE = os.path.join(app.config['DATA_FOLDER'], 'orders.csv')
STOCKS_FILE = os.path.join(app.config['DATA_FOLDER'], 'stocks.json')
PROCESSED_ORDERS_FILE = os.path.join(app.config['DATA_FOLDER'], 'processed_orders.json')

# Создаем файлы данных, если их нет
if not os.path.exists(PRODUCTS_FILE):
    with open(PRODUCTS_FILE, 'w', encoding='utf-8') as f:
        f.write("Название,Артикул,Модель\n")

if not os.path.exists(ORDERS_FILE):
    with open(ORDERS_FILE, 'w', encoding='utf-8') as f:
        f.write("Operation Type,Date,Order Number,Product Name,Product Article,Quantity\n")

if not os.path.exists(STOCKS_FILE):
    with open(STOCKS_FILE, 'w', encoding='utf-8') as f:
        json.dump({}, f)

if not os.path.exists(PROCESSED_ORDERS_FILE):
    with open(PROCESSED_ORDERS_FILE, 'w', encoding='utf-8') as f:
        json.dump([], f)


# ==================== КЭШИРОВАНИЕ ДАННЫХ ====================

class DataCache:
    """Кэш для данных с блокировками"""
    def __init__(self):
        self.cache = {}
        self.locks = {}
        self.max_age = 300  # 5 минут
    
    def get(self, key, loader):
        if key not in self.locks:
            self.locks[key] = threading.Lock()
        
        with self.locks[key]:
            if key in self.cache:
                data, timestamp = self.cache[key]
                if time.time() - timestamp < self.max_age:
                    return data
            
            # Загружаем данные
            data = loader()
            self.cache[key] = (data, time.time())
            return data
    
    def invalidate(self, key):
        if key in self.cache:
            del self.cache[key]

data_cache = DataCache()


# ==================== ИНДЕКС ДЛЯ ПОИСКА ТОВАРОВ ====================

class ProductIndex:
    """Класс для быстрого поиска товаров"""
    def __init__(self, products):
        self.by_article = {p['article']: p for p in products}
        self.by_name_lower = {p['name'].lower(): p for p in products}
        self.prefix_index = self._build_prefix_index(products)
        self.last_update = datetime.now()
    
    def _build_prefix_index(self, products):
        prefix_index = {}
        for p in products:
            if '_' in p['article']:
                parts = p['article'].split('_')
                for i in range(1, len(parts) + 1):
                    prefix = '_'.join(parts[:i])
                    if prefix not in prefix_index:
                        prefix_index[prefix] = []
                    prefix_index[prefix].append(p)
        return prefix_index
    
    def search(self, query):
        """Поиск по различным полям"""
        results = []
        query_lower = query.lower()
        
        # Поиск по артикулу
        if query in self.by_article:
            results.append(self.by_article[query])
        
        # Поиск по имени
        if query_lower in self.by_name_lower:
            results.append(self.by_name_lower[query_lower])
        
        # Поиск по префиксу
        if query in self.prefix_index:
            results.extend(self.prefix_index[query])
        
        # Удаляем дубликаты
        seen = set()
        unique_results = []
        for p in results:
            if p['article'] not in seen:
                seen.add(p['article'])
                unique_results.append(p)
        
        return unique_results
    
    def search_by_part(self, part):
        """Поиск по части артикула"""
        results = []
        part_lower = part.lower()
        
        for product in self.by_article.values():
            article = product['article'].lower()
            name = product['name'].lower()
            model = product.get('model', '').lower()
            
            if part_lower in article or part_lower in name or part_lower in model:
                results.append(product)
        
        return results


# ==================== АСИНХРОННАЯ ОБРАБОТКА ФАЙЛОВ ====================

class AsyncFileProcessor:
    """Асинхронная обработка файлов"""
    def __init__(self):
        self.task_queue = queue.Queue()
        self.results = {}
        self.thread = threading.Thread(target=self._worker, daemon=True)
        self.thread.start()
    
    def _worker(self):
        while True:
            task_id, file_path, processor = self.task_queue.get()
            try:
                result = processor(file_path)
                self.results[task_id] = {
                    'status': 'completed',
                    'result': result,
                    'time': datetime.now()
                }
            except Exception as e:
                self.results[task_id] = {
                    'status': 'error',
                    'error': str(e),
                    'time': datetime.now()
                }
            finally:
                self.task_queue.task_done()
    
    def submit(self, file_path, processor):
        task_id = str(uuid.uuid4())
        self.task_queue.put((task_id, file_path, processor))
        return task_id
    
    def get_result(self, task_id):
        return self.results.get(task_id)

file_processor = AsyncFileProcessor()


# ==================== СИСТЕМНЫЕ МЕТРИКИ ====================

class SystemMetrics:
    """Сбор метрик системы"""
    def __init__(self):
        self.operation_counter = Counter()
        self.response_times = defaultdict(list)
        self.start_time = datetime.now()
    
    def record_operation(self, operation_type):
        self.operation_counter[operation_type] += 1
    
    def record_response_time(self, endpoint, time_ms):
        self.response_times[endpoint].append(time_ms)
        # Оставляем только последние 100 замеров
        if len(self.response_times[endpoint]) > 100:
            self.response_times[endpoint].pop(0)
    
    def get_stats(self):
        return {
            'uptime': str(datetime.now() - self.start_time),
            'operations': dict(self.operation_counter),
            'avg_response_times': {
                endpoint: sum(times)/len(times) if times else 0
                for endpoint, times in self.response_times.items()
            },
            'system': {
                'cpu_percent': psutil.cpu_percent(),
                'memory_percent': psutil.virtual_memory().percent,
                'disk_usage': psutil.disk_usage('/').percent
            }
        }

metrics = SystemMetrics()

@app.before_request
def before_request():
    request.start_time = time.time()

@app.after_request
def after_request(response):
    if hasattr(request, 'start_time'):
        elapsed = (time.time() - request.start_time) * 1000
        if request.endpoint:
            metrics.record_response_time(request.endpoint, elapsed)
    return response


# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================

def login_required(f):
    """Декоратор для проверки авторизации"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('Пожалуйста, войдите в систему', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


@app.context_processor
def utility_processor():
    def now():
        return datetime.now()
    return dict(now=now)


def load_products():
    """Загружает список товаров из CSV"""
    products = []
    try:
        if os.path.exists(PRODUCTS_FILE) and os.path.getsize(PRODUCTS_FILE) > 0:
            with open(PRODUCTS_FILE, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader, None)  # Пропускаем заголовок
                for row in reader:
                    if len(row) >= 2:
                        product = {
                            'name': row[0].strip(),
                            'article': row[1].strip()
                        }
                        if len(row) >= 3:
                            product['model'] = row[2].strip()
                        products.append(product)
        app.logger.info(f"Загружено {len(products)} товаров")
    except Exception as e:
        app.logger.error(f"Ошибка загрузки товаров: {e}")
    return products


def save_products(products):
    """Сохраняет список товаров в CSV"""
    try:
        with open(PRODUCTS_FILE, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Название', 'Артикул', 'Модель'])
            for product in products:
                writer.writerow([
                    product['name'], 
                    product['article'], 
                    product.get('model', '')
                ])
        app.logger.info(f"Сохранено {len(products)} товаров")
        data_cache.invalidate('products')
        return True
    except Exception as e:
        app.logger.error(f"Ошибка сохранения товаров: {e}")
        return False


def load_stocks():
    """Загружает остатки товаров из JSON"""
    try:
        if os.path.exists(STOCKS_FILE) and os.path.getsize(STOCKS_FILE) > 0:
            with open(STOCKS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        app.logger.error(f"Ошибка загрузки остатков: {e}")
        return {}


def save_stocks(stocks):
    """Сохраняет остатки товаров в JSON"""
    try:
        with open(STOCKS_FILE, 'w', encoding='utf-8') as f:
            json.dump(stocks, f, indent=2, ensure_ascii=False)
        app.logger.info("Остатки сохранены")
        data_cache.invalidate('stocks')
        return True
    except Exception as e:
        app.logger.error(f"Ошибка сохранения остатков: {e}")
        return False


def load_orders():
    """Загружает все заказы из CSV"""
    orders = []
    try:
        if os.path.exists(ORDERS_FILE) and os.path.getsize(ORDERS_FILE) > 0:
            with open(ORDERS_FILE, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    order = {}
                    for key, value in row.items():
                        if key == 'Quantity':
                            try:
                                order[key] = int(value) if value and value.strip() else 0
                            except:
                                order[key] = 0
                        else:
                            order[key] = str(value) if value else ''
                    orders.append(order)
            app.logger.info(f"Загружено {len(orders)} заказов")
    except Exception as e:
        app.logger.error(f"Ошибка загрузки заказов: {e}")
    return orders


def save_order(order_data):
    """Сохраняет новый заказ в CSV"""
    try:
        file_exists = os.path.exists(ORDERS_FILE)
        with open(ORDERS_FILE, 'a', newline='', encoding='utf-8') as f:
            fieldnames = ["Operation Type", "Date", "Order Number", "Product Name", "Product Article", "Quantity"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            if not file_exists or os.path.getsize(ORDERS_FILE) == 0:
                writer.writeheader()
            
            writer.writerow(order_data)
            f.flush()
            
        app.logger.info(f"Заказ сохранен: {order_data}")
        metrics.record_operation('order_saved')
        return True
    except Exception as e:
        app.logger.error(f"Ошибка при сохранении заказа: {e}")
        return False


def save_all_orders(orders):
    """Сохраняет все заказы в CSV"""
    try:
        with open(ORDERS_FILE, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ["Operation Type", "Date", "Order Number", "Product Name", "Product Article", "Quantity"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for order in orders:
                writer.writerow(order)
        app.logger.info(f"Сохранено {len(orders)} заказов")
        data_cache.invalidate('orders')
        return True
    except Exception as e:
        app.logger.error(f"Ошибка при сохранении заказов: {e}")
        return False


def load_processed_orders():
    """Загружает список обработанных заказов"""
    try:
        if os.path.exists(PROCESSED_ORDERS_FILE) and os.path.getsize(PROCESSED_ORDERS_FILE) > 0:
            with open(PROCESSED_ORDERS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    except Exception as e:
        app.logger.error(f"Ошибка загрузки обработанных заказов: {e}")
        return []


def save_processed_order(order_number):
    """Сохраняет номер обработанного заказа"""
    try:
        processed = load_processed_orders()
        if order_number not in processed:
            processed.append(order_number)
            with open(PROCESSED_ORDERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(processed, f, indent=2)
            app.logger.info(f"Заказ {order_number} помечен как обработанный")
        return True
    except Exception as e:
        app.logger.error(f"Ошибка при сохранении обработанного заказа: {e}")
        return False


def validate_article(article):
    """
    Расширенная валидация артикулов
    """
    if not article or not article.strip():
        return False, "Артикул не может быть пустым"
    
    article = article.strip()
    
    if len(article) > 100:
        return False, "Артикул слишком длинный (максимум 100 символов)"
    
    # Проверка на недопустимые символы
    invalid_chars = set(article) - set('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyzАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯабвгдеёжзийклмнопрстуфхцчшщъыьэюя0123456789-._/:')
    
    if invalid_chars:
        return False, f"Артикул содержит недопустимые символы: {', '.join(invalid_chars)}"
    
    # Проверка на последовательность из одинаковых символов
    if re.search(r'(.)\1{4,}', article):
        return False, "Артикул содержит слишком много одинаковых символов подряд"
    
    return True, "OK"


def backup_data():
    """Создает резервную копию всех данных"""
    try:
        backup_dir = os.path.join(app.config['DATA_FOLDER'], 'backups')
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(backup_dir, f'backup_{timestamp}.zip')
        
        with zipfile.ZipFile(backup_file, 'w') as zipf:
            for file in [PRODUCTS_FILE, ORDERS_FILE, STOCKS_FILE, PROCESSED_ORDERS_FILE]:
                if os.path.exists(file):
                    zipf.write(file, os.path.basename(file))
        
        # Удаляем старые бэкапы (оставляем последние 10)
        backups = sorted([f for f in os.listdir(backup_dir) if f.startswith('backup_')])
        for old_backup in backups[:-10]:
            os.remove(os.path.join(backup_dir, old_backup))
        
        app.logger.info(f"Создана резервная копия: {backup_file}")
        return True, backup_file
    except Exception as e:
        app.logger.error(f"Ошибка при создании бэкапа: {e}")
        return False, str(e)


def clean_dataframe(df):
    """Очищает DataFrame от полностью пустых строк"""
    if df.empty:
        return df
    
    # Удаляем строки, где все значения NaN
    df = df.dropna(how='all')
    
    # Удаляем строки, где пустые обязательные поля (первые две колонки)
    if len(df.columns) >= 2:
        col1 = df.columns[0]
        col2 = df.columns[1]
        df = df.dropna(subset=[col1, col2], how='all')
    
    return df


def detect_encoding(filepath):
    """Определяет кодировку файла"""
    try:
        with open(filepath, 'rb') as f:
            raw_data = f.read(10000)
            result = chardet.detect(raw_data)
            return result['encoding'] or 'utf-8'
    except:
        return 'utf-8'


# ==================== ФУНКЦИИ ДЛЯ РАБОТЫ С СОСТАВНЫМИ АРТИКУЛАМИ ====================

def find_matching_products_enhanced(composite_article, products):
    """
    Улучшенный поиск с учетом похожих артикулов
    """
    if not composite_article:
        return [], []
    
    found_products = []
    not_found = []
    
    # Создаем словарь для быстрого поиска
    products_dict = {p['article']: p for p in products}
    
    # Также создаем словарь для поиска по началу артикула
    prefix_dict = {}
    for p in products:
        article = p['article']
        # Сохраняем первые части до _
        if '_' in article:
            prefix = article.split('_')[0]
            if prefix not in prefix_dict:
                prefix_dict[prefix] = []
            prefix_dict[prefix].append(p)
    
    app.logger.info(f"ENHANCED поиск в артикуле: '{composite_article}'")
    
    def recursive_search(search_str, found_list):
        if not search_str or len(search_str.strip()) == 0:
            return
        
        app.logger.info(f"  Поиск в: '{search_str}'")
        
        # 1. Проверяем точное совпадение
        if search_str in products_dict:
            app.logger.info(f"    -> Найдено точное совпадение: '{search_str}'")
            found_list.append(products_dict[search_str])
            return
        
        # 2. Если есть _, пробуем найти самое длинное совпадение
        if '_' in search_str:
            parts = search_str.split('_')
            
            # Пробуем от самого длинного к короткому
            for i in range(len(parts), 0, -1):
                candidate = '_'.join(parts[:i])
                if candidate in products_dict:
                    app.logger.info(f"    -> Найдено совпадение: '{candidate}'")
                    found_list.append(products_dict[candidate])
                    
                    # Ищем в оставшейся части
                    remaining = search_str[len(candidate):].lstrip('_')
                    if remaining:
                        app.logger.info(f"    -> Ищем в остатке: '{remaining}'")
                        recursive_search(remaining, found_list)
                    return
            
            # 3. Пробуем найти по префиксу (первая часть)
            first_part = parts[0]
            if first_part in prefix_dict:
                app.logger.info(f"    -> Найдены товары с префиксом '{first_part}': {len(prefix_dict[first_part])}")
                # Берем первый подходящий
                for product in prefix_dict[first_part]:
                    app.logger.info(f"    -> Используем: {product['article']}")
                    found_list.append(product)
                    
                    # Остаток - все что после первой части
                    remaining = '_'.join(parts[1:])
                    if remaining:
                        app.logger.info(f"    -> Ищем в остатке: '{remaining}'")
                        recursive_search(remaining, found_list)
                    return
            
            # 4. Пробуем найти любой из существующих артикулов внутри строки
            for article in products_dict.keys():
                if article in search_str and len(article) > 3:
                    app.logger.info(f"    -> Найдено вхождение: '{article}'")
                    
                    # Находим позицию
                    pos = search_str.find(article)
                    
                    # Часть до
                    before = search_str[:pos].rstrip('_')
                    if before:
                        app.logger.info(f"    -> Часть ДО: '{before}'")
                        recursive_search(before, found_list)
                    
                    # Добавляем найденный
                    found_list.append(products_dict[article])
                    
                    # Часть после
                    after = search_str[pos + len(article):].lstrip('_')
                    if after:
                        app.logger.info(f"    -> Часть ПОСЛЕ: '{after}'")
                        recursive_search(after, found_list)
                    return
        
        # 5. Если ничего не нашли
        app.logger.info(f"    -> НИЧЕГО не найдено в '{search_str}'")
        not_found.append(search_str)
    
    recursive_search(composite_article, found_products)
    
    # Удаляем дубликаты
    unique_found = []
    seen = set()
    for p in found_products:
        if p['article'] not in seen:
            seen.add(p['article'])
            unique_found.append(p)
    
    return unique_found, not_found


def process_composite_article(composite_article, products, stocks, order_number, 
                            quantity, product_name, model, new_products_count_ref=None, 
                            missing_products=None):
    """
    Обрабатывает составной артикул, находя все возможные товары
    """
    app.logger.info(f"Обработка составного артикула: {composite_article}")
    
    # Используем улучшенную функцию поиска
    found_products, not_found_parts = find_matching_products_enhanced(composite_article, products)
    
    app.logger.info(f"  Найдено товаров: {len(found_products)}")
    app.logger.info(f"  Ненайденные части: {not_found_parts}")
    
    # Счетчики для возврата
    found_count = len(found_products)
    created_count = 0
    
    # Обрабатываем найденные товары
    for product in found_products:
        order_data = {
            "Operation Type": "Расход",
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Order Number": order_number,
            "Product Name": product['name'],
            "Product Article": product['article'],
            "Quantity": quantity
        }
        
        if save_order(order_data):
            app.logger.info(f"    Заказ сохранен для арт. {product['article']}")
        else:
            app.logger.error(f"    Ошибка сохранения заказа для арт. {product['article']}")
        
        current = stocks.get(product['article'], 0)
        stocks[product['article']] = current - quantity
        app.logger.info(f"    Остаток обновлен: {current} -> {stocks[product['article']]}")
    
    # Создаем новые товары для ненайденных частей
    for part in not_found_parts:
        if part and len(part.strip()) > 0:
            # Проверяем, не существует ли уже товар с таким артикулом
            existing = False
            for p in products:
                if p['article'] == part:
                    existing = True
                    app.logger.info(f"    Часть '{part}' найдена при повторной проверке!")
                    
                    # Сохраняем заказ для существующего товара
                    order_data = {
                        "Operation Type": "Расход",
                        "Date": datetime.now().strftime("%Y-%m-%d"),
                        "Order Number": order_number,
                        "Product Name": p['name'],
                        "Product Article": part,
                        "Quantity": quantity
                    }
                    if save_order(order_data):
                        app.logger.info(f"    Заказ сохранен для существующего арт. {part}")
                    
                    current = stocks.get(part, 0)
                    stocks[part] = current - quantity
                    found_count += 1  # Увеличиваем счетчик найденных
                    break
            
            if not existing:
                # Создаем новый товар
                new_name = model or product_name or f"Товар из комплекта {part}"
                
                new_product = {
                    'name': new_name,
                    'article': part,
                    'model': model if model else ''
                }
                products.append(new_product)
                created_count += 1
                if new_products_count_ref is not None:
                    new_products_count_ref['count'] += 1
                
                app.logger.info(f"    >>> СОЗДАН НОВЫЙ ТОВАР: {new_name} (арт. {part})")
                
                # Сохраняем заказ для нового товара
                order_data = {
                    "Operation Type": "Расход",
                    "Date": datetime.now().strftime("%Y-%m-%d"),
                    "Order Number": order_number,
                    "Product Name": new_name,
                    "Product Article": part,
                    "Quantity": quantity
                }
                if save_order(order_data):
                    app.logger.info(f"    Заказ сохранен для нового арт. {part}")
                
                # Устанавливаем отрицательный остаток (расход)
                stocks[part] = -quantity
    
    return found_count, created_count


def process_single_article(article, model, product_name, quantity, products, stocks, 
                          order_number, display_name=None, new_products_count_ref=None, 
                          missing_products=None):
    """Обрабатывает один артикул товара"""
    
    app.logger.info(f"Обработка одиночного артикула: {article}, модель: {model}, название: {product_name}")
    
    # Поиск товара по артикулу
    product_exists = False
    existing_product = None
    
    for product in products:
        if product['article'] == article:
            product_exists = True
            existing_product = product
            app.logger.info(f"    Найден товар по артикулу: {product['name']}")
            break
    
    # Если не нашли по артикулу, ищем по модели
    if not product_exists and model:
        for product in products:
            if product.get('model') == model or product['name'] == model:
                product_exists = True
                existing_product = product
                article = product['article']  # Обновляем артикул
                app.logger.info(f"    Найден товар по модели: {product['name']}")
                break
    
    # Если не нашли по модели, ищем по названию
    if not product_exists and product_name:
        for product in products:
            if product['name'] == product_name:
                product_exists = True
                existing_product = product
                article = product['article']  # Обновляем артикул
                if not model:
                    model = product.get('model', '')
                app.logger.info(f"    Найден товар по названию: {product['name']}")
                break
    
    # Если не нашли, создаем новый товар
    if not product_exists:
        if not article:
            # Генерируем артикул из названия или модели
            if model:
                base = re.sub(r'[^A-Za-z0-9]', '', model)[:8]
            elif product_name:
                base = re.sub(r'[^A-Za-z0-9]', '', product_name)[:8]
            else:
                base = "NEW"
            
            # Проверяем уникальность
            counter = 1
            new_article = base
            while any(p['article'] == new_article for p in products):
                new_article = f"{base}_{counter}"
                counter += 1
            article = new_article
        
        # Создаем новый товар
        new_product = {
            'name': display_name or model or product_name or f"Товар {article}", 
            'article': article
        }
        if model:
            new_product['model'] = model
        
        products.append(new_product)
        
        if new_products_count_ref is not None:
            new_products_count_ref['count'] += 1
        
        app.logger.info(f"    >>> СОЗДАН НОВЫЙ ТОВАР: {new_product['name']} (арт. {article})")
        existing_product = new_product
        product_exists = True
    else:
        app.logger.info(f"    Используется существующий товар: {existing_product['name']} (арт. {article})")
    
    # Сохраняем операцию расхода
    if product_exists and existing_product:
        order_data = {
            "Operation Type": "Расход",
            "Date": datetime.now().strftime("%Y-%m-%d"),
            "Order Number": order_number,
            "Product Name": existing_product['name'],
            "Product Article": article,
            "Quantity": quantity
        }
        
        if save_order(order_data):
            app.logger.info(f"    Заказ сохранен для арт. {article}")
        else:
            app.logger.error(f"    Ошибка сохранения заказа для арт. {article}")
            if missing_products is not None:
                missing_products.append(f"{article} - ошибка сохранения")
        
        current = stocks.get(article, 0)
        stocks[article] = current - quantity
        app.logger.info(f"    Остаток обновлен: {current} -> {stocks[article]}")
        
        return True
    else:
        app.logger.error(f"    Не удалось создать или найти товар для арт. {article}")
        if missing_products is not None:
            missing_products.append(f"{article} - не удалось создать товар")
        return False


# ==================== МАРШРУТЫ АВТОРИЗАЦИИ ====================

@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if username == 'admin' and password_hash == hashlib.sha256('admin'.encode()).hexdigest():
            session['user'] = username
            session.permanent = True
            app.logger.info(f"Пользователь {username} вошел в систему")
            flash('Вход выполнен успешно', 'success')
            return redirect(url_for('dashboard'))
        else:
            app.logger.warning(f"Неудачная попытка входа: {username}")
            flash('Неверное имя пользователя или пароль', 'danger')
    
    return render_template('login.html')


@app.route('/logout')
def logout():
    user = session.get('user')
    session.pop('user', None)
    app.logger.info(f"Пользователь {user} вышел из системы")
    flash('Вы вышли из системы', 'info')
    return redirect(url_for('login'))


# ==================== МАРШРУТЫ ДАШБОРДА ====================

@app.route('/dashboard')
@login_required
def dashboard():
    products = data_cache.get('products', load_products)
    stocks = data_cache.get('stocks', load_stocks)
    orders = data_cache.get('orders', load_orders)
    
    for product in products:
        product['stock'] = stocks.get(product['article'], 0)
    
    total_products = len(products)
    total_orders = len(set([o.get('Order Number', '') for o in orders if o.get('Order Number')]))
    total_stock_items = sum(stocks.values())
    low_stock_count = sum(1 for p in products if stocks.get(p['article'], 0) < 10)
    
    recent_orders = sorted(orders, key=lambda x: x.get('Date', ''), reverse=True)[:10]
    
    return render_template('dashboard.html',
                         total_products=total_products,
                         total_orders=total_orders,
                         total_stock_items=total_stock_items,
                         low_stock_count=low_stock_count,
                         recent_orders=recent_orders,
                         products=products)


# ==================== МАРШРУТЫ ТОВАРОВ ====================

@app.route('/products')
@login_required
def products_list():
    products = data_cache.get('products', load_products)
    stocks = data_cache.get('stocks', load_stocks)
    
    for product in products:
        product['stock'] = stocks.get(product['article'], 0)
    
    return render_template('products.html', products=products)


@app.route('/products/add', methods=['POST'])
@login_required
def add_product():
    try:
        data = request.json
        name = data.get('name', '').strip()
        article = data.get('article', '').strip()
        model = data.get('model', '').strip()
        
        if not name or not article:
            return jsonify({'success': False, 'error': 'Не все поля заполнены'})
        
        is_valid, error_msg = validate_article(article)
        if not is_valid:
            return jsonify({'success': False, 'error': error_msg})
        
        products = load_products()
        
        for product in products:
            if product['article'] == article:
                return jsonify({'success': False, 'error': 'Товар с таким артикулом уже существует'})
        
        products.append({'name': name, 'article': article, 'model': model})
        
        if save_products(products):
            metrics.record_operation('product_added')
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Ошибка при сохранении'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/delete/<article>', methods=['POST'])
@login_required
def delete_product(article):
    try:
        products = load_products()
        stocks = load_stocks()
        
        products = [p for p in products if p['article'] != article]
        
        if article in stocks:
            del stocks[article]
            save_stocks(stocks)
        
        if save_products(products):
            metrics.record_operation('product_deleted')
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Ошибка при удалении'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/edit', methods=['POST'])
@login_required
def edit_product():
    try:
        data = request.json
        old_article = data.get('old_article')
        new_name = data.get('name', '').strip()
        new_article = data.get('article', '').strip()
        new_model = data.get('model', '').strip()
        
        if not old_article or not new_name or not new_article:
            return jsonify({'success': False, 'error': 'Не все поля заполнены'})
        
        is_valid, error_msg = validate_article(new_article)
        if not is_valid:
            return jsonify({'success': False, 'error': error_msg})
        
        products = load_products()
        stocks = load_stocks()
        
        if old_article != new_article:
            for product in products:
                if product['article'] == new_article:
                    return jsonify({'success': False, 'error': 'Товар с таким артикулом уже существует'})
        
        product_found = False
        for product in products:
            if product['article'] == old_article:
                product['name'] = new_name
                product['article'] = new_article
                product['model'] = new_model
                product_found = True
                break
        
        if not product_found:
            return jsonify({'success': False, 'error': 'Товар не найден'})
        
        if old_article != new_article and old_article in stocks:
            stocks[new_article] = stocks.pop(old_article)
            save_stocks(stocks)
        
        if save_products(products):
            metrics.record_operation('product_edited')
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Ошибка при сохранении'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/search')
@login_required
def search_products_api():
    try:
        query = request.args.get('q', '').lower()
        products = load_products()
        stocks = load_stocks()
        
        results = []
        for product in products:
            if (query in product['name'].lower() or 
                query in product['article'].lower() or 
                (product.get('model') and query in product['model'].lower())):
                results.append({
                    'name': product['name'],
                    'article': product['article'],
                    'model': product.get('model', ''),
                    'stock': stocks.get(product['article'], 0)
                })
        
        return jsonify(results)
    except Exception as e:
        return jsonify([])


@app.route('/products/search/exact')
@login_required
def search_products_exact():
    try:
        article = request.args.get('article', '').strip()
        products = load_products()
        stocks = load_stocks()
        
        for product in products:
            if product['article'] == article:
                return jsonify({
                    'found': True,
                    'product': {
                        'name': product['name'],
                        'article': product['article'],
                        'stock': stocks.get(article, 0)
                    }
                })
        
        return jsonify({'found': False})
    except Exception as e:
        return jsonify({'found': False, 'error': str(e)})


@app.route('/products/search/by-part', methods=['GET'])
@login_required
def search_products_by_part():
    """Поиск товаров по части артикула (например, 18.102wh)"""
    try:
        part = request.args.get('part', '').strip().lower()
        if not part or len(part) < 3:
            return jsonify({'success': False, 'error': 'Введите минимум 3 символа'})
        
        products = load_products()
        stocks = load_stocks()
        
        results = []
        for product in products:
            article = product['article'].lower()
            name = product['name'].lower()
            model = product.get('model', '').lower()
            
            # Ищем вхождение части в артикуле, названии или модели
            if part in article or part in name or part in model:
                # Определяем, является ли артикул составным
                is_composite = '_' in product['article']
                
                # Находим позицию вхождения для подсветки
                position = article.find(part) if part in article else -1
                
                results.append({
                    'name': product['name'],
                    'article': product['article'],
                    'model': product.get('model', ''),
                    'stock': stocks.get(product['article'], 0),
                    'is_composite': is_composite,
                    'match_position': position,
                    'match_field': 'article' if part in article else ('name' if part in name else 'model')
                })
        
        # Сортируем: сначала точные совпадения, потом по позиции
        results.sort(key=lambda x: (
            - (1 if x['match_position'] == 0 else 0),  # Сначала те, что начинаются с искомого
            x['match_position'] if x['match_position'] >= 0 else 999,  # Потом по позиции
            x['article']  # Затем по алфавиту
        ))
        
        return jsonify({
            'success': True,
            'part': part,
            'count': len(results),
            'results': results[:100]  # Ограничим до 100 результатов
        })
        
    except Exception as e:
        app.logger.error(f"Ошибка при поиске по части артикула: {e}")
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ МАССОВОЙ ЗАГРУЗКИ ТОВАРОВ ====================

@app.route('/products/upload', methods=['GET', 'POST'])
@login_required
def upload_products():
    if request.method == 'GET':
        return render_template('upload_products.html')
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in {'.xlsx', '.xls', '.csv'}:
            return jsonify({'success': False, 'error': 'Неверный формат файла. Разрешены: Excel, CSV'})
        
        filename = secure_filename(f"upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            df = None
            
            if file_ext == '.csv':
                encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'windows-1251', 'koi8-r', 'latin1', 'iso-8859-1', 'cp866', 'mac-cyrillic']
                separators = [',', ';', '\t', '|', ' ']
                
                for encoding in encodings:
                    for sep in separators:
                        try:
                            test_df = pd.read_csv(filepath, encoding=encoding, sep=sep, nrows=5)
                            if len(test_df.columns) >= 2:
                                df = pd.read_csv(filepath, encoding=encoding, sep=sep)
                                break
                        except:
                            continue
                    if df is not None:
                        break
                
                if df is None:
                    encoding = detect_encoding(filepath)
                    for sep in separators:
                        try:
                            df = pd.read_csv(filepath, encoding=encoding, sep=sep)
                            if len(df.columns) >= 2:
                                break
                        except:
                            continue
                
                if df is None:
                    return jsonify({'success': False, 'error': 'Не удалось прочитать CSV файл'})
            else:
                df = pd.read_excel(filepath)
            
            df = clean_dataframe(df)
            
            if df.empty:
                return jsonify({'success': False, 'error': 'Файл не содержит данных'})
            
            df.columns = [str(col).strip() for col in df.columns]
            
            column_mapping = {}
            
            for col in df.columns:
                col_lower = col.lower().strip()
                
                if any(x in col_lower for x in ['название', 'товар', 'name', 'product', 'наименование']):
                    column_mapping['name'] = col
                elif any(x in col_lower for x in ['артикул', 'article', 'код', 'sku', 'vendor']):
                    column_mapping['article'] = col
                elif any(x in col_lower for x in ['модель', 'model']):
                    column_mapping['model'] = col
                elif any(x in col_lower for x in ['цена', 'price', 'cost']):
                    column_mapping['price'] = col
                elif any(x in col_lower for x in ['остаток', 'stock', 'quantity']):
                    column_mapping['stock'] = col
                elif any(x in col_lower for x in ['категория', 'category']):
                    column_mapping['category'] = col
                elif any(x in col_lower for x in ['описание', 'description']):
                    column_mapping['description'] = col
            
            if 'name' not in column_mapping or 'article' not in column_mapping:
                return jsonify({'success': False, 'error': 'Не найдены обязательные колонки "Название" и "Артикул"'})
            
            products = load_products()
            existing_articles = {p['article']: p for p in products}
            
            new_products = []
            updated_products = []
            errors = []
            preview_data = []
            skipped_empty = 0
            
            for idx, row in df.iterrows():
                try:
                    name_val = row[column_mapping['name']]
                    article_val = row[column_mapping['article']]
                    
                    if pd.isna(name_val) and pd.isna(article_val):
                        skipped_empty += 1
                        continue
                    
                    name = str(name_val).strip() if not pd.isna(name_val) else ''
                    article = str(article_val).strip() if not pd.isna(article_val) else ''
                    
                    if not article:
                        errors.append(f"Строка {idx + 2}: отсутствует артикул")
                        continue
                    
                    if not name:
                        name = f"Товар {article}"
                    
                    is_valid, error_msg = validate_article(article)
                    if not is_valid:
                        errors.append(f"Строка {idx + 2}: {error_msg}")
                        continue
                    
                    model = ''
                    if 'model' in column_mapping and not pd.isna(row[column_mapping['model']]):
                        model = str(row[column_mapping['model']]).strip()
                    
                    price = None
                    if 'price' in column_mapping and not pd.isna(row[column_mapping['price']]):
                        try:
                            price = float(row[column_mapping['price']])
                        except:
                            pass
                    
                    stock = None
                    if 'stock' in column_mapping and not pd.isna(row[column_mapping['stock']]):
                        try:
                            stock = int(float(row[column_mapping['stock']]))
                        except:
                            pass
                    
                    category = ''
                    if 'category' in column_mapping and not pd.isna(row[column_mapping['category']]):
                        category = str(row[column_mapping['category']]).strip()
                    
                    description = ''
                    if 'description' in column_mapping and not pd.isna(row[column_mapping['description']]):
                        description = str(row[column_mapping['description']]).strip()
                    
                    product_data = {
                        'name': name,
                        'article': article,
                        'model': model,
                        'price': price,
                        'stock': stock,
                        'category': category,
                        'description': description
                    }
                    
                    preview_data.append(product_data)
                    
                    if article in existing_articles:
                        existing = existing_articles[article]
                        if name:
                            existing['name'] = name
                        if model:
                            existing['model'] = model
                        updated_products.append(article)
                    else:
                        new_product = {'name': name, 'article': article}
                        if model:
                            new_product['model'] = model
                        products.append(new_product)
                        existing_articles[article] = new_product
                        new_products.append(article)
                    
                except Exception as e:
                    errors.append(f"Строка {idx + 2}: {str(e)}")
            
            session['upload_preview'] = {
                'products': preview_data,
                'new_count': len(new_products),
                'updated_count': len(updated_products)
            }
            
            if os.path.exists(filepath):
                os.remove(filepath)
            
            return jsonify({
                'success': True,
                'preview': preview_data[:50],
                'stats': {
                    'total': len(preview_data),
                    'new': len(new_products),
                    'updated': len(updated_products),
                    'errors': len(errors),
                    'skipped': skipped_empty
                },
                'errors': errors[:10]
            })
            
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'success': False, 'error': str(e)})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/upload/new', methods=['GET', 'POST'])
@login_required
def upload_new_products():
    if request.method == 'GET':
        return render_template('upload_new_products.html')
    
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in {'.xlsx', '.xls', '.csv'}:
            return jsonify({'success': False, 'error': 'Неверный формат файла. Разрешены: Excel, CSV'})
        
        filename = secure_filename(f"new_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            if file_ext == '.csv':
                encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'windows-1251', 'koi8-r']
                df = None
                for encoding in encodings:
                    try:
                        df = pd.read_csv(filepath, encoding=encoding)
                        break
                    except:
                        continue
                if df is None:
                    encoding = detect_encoding(filepath)
                    df = pd.read_csv(filepath, encoding=encoding)
            else:
                df = pd.read_excel(filepath)
            
            df = clean_dataframe(df)
            
            if df.empty:
                return jsonify({'success': False, 'error': 'Файл не содержит данных'})
            
            df.columns = [str(col).strip() for col in df.columns]
            
            column_mapping = {}
            
            for col in df.columns:
                col_lower = col.lower().strip()
                
                if any(x in col_lower for x in ['название', 'товар', 'name', 'product']):
                    column_mapping['name'] = col
                elif any(x in col_lower for x in ['артикул', 'article', 'код', 'sku']):
                    column_mapping['article'] = col
                elif any(x in col_lower for x in ['модель', 'model']):
                    column_mapping['model'] = col
                elif any(x in col_lower for x in ['цена', 'price']):
                    column_mapping['price'] = col
                elif any(x in col_lower for x in ['остаток', 'stock']):
                    column_mapping['stock'] = col
            
            if 'name' not in column_mapping or 'article' not in column_mapping:
                return jsonify({'success': False, 'error': 'Не найдены обязательные колонки "Название" и "Артикул"'})
            
            products = load_products()
            existing_articles = {p['article']: p for p in products}
            
            new_products = []
            duplicates = []
            errors = []
            preview_data = []
            
            for idx, row in df.iterrows():
                try:
                    name = str(row[column_mapping['name']]).strip() if not pd.isna(row[column_mapping['name']]) else ''
                    article = str(row[column_mapping['article']]).strip() if not pd.isna(row[column_mapping['article']]) else ''
                    
                    if not article:
                        continue
                    
                    is_valid, error_msg = validate_article(article)
                    if not is_valid:
                        errors.append(f"Строка {idx + 2}: {error_msg}")
                        continue
                    
                    if article in existing_articles:
                        duplicates.append({
                            'row': idx + 2,
                            'article': article,
                            'existing_name': existing_articles[article]['name'],
                            'new_name': name
                        })
                        continue
                    
                    model = ''
                    if 'model' in column_mapping and not pd.isna(row[column_mapping['model']]):
                        model = str(row[column_mapping['model']]).strip()
                    
                    price = None
                    if 'price' in column_mapping and not pd.isna(row[column_mapping['price']]):
                        try:
                            price = float(row[column_mapping['price']])
                        except:
                            pass
                    
                    stock = None
                    if 'stock' in column_mapping and not pd.isna(row[column_mapping['stock']]):
                        try:
                            stock = int(float(row[column_mapping['stock']]))
                        except:
                            pass
                    
                    product_data = {
                        'name': name or f"Товар {article}",
                        'article': article,
                        'model': model,
                        'price': price,
                        'stock': stock
                    }
                    
                    preview_data.append(product_data)
                    new_products.append(product_data)
                    
                except Exception as e:
                    errors.append(f"Строка {idx + 2}: {str(e)}")
            
            session['new_products_preview'] = {
                'products': new_products,
                'count': len(new_products),
                'duplicates': duplicates
            }
            
            if os.path.exists(filepath):
                os.remove(filepath)
            
            return jsonify({
                'success': True,
                'preview': new_products[:50],
                'stats': {
                    'total': len(new_products),
                    'duplicates': len(duplicates),
                    'errors': len(errors)
                },
                'duplicates': duplicates[:10],
                'errors': errors[:10]
            })
            
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'success': False, 'error': str(e)})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/upload/confirm', methods=['POST'])
@login_required
def confirm_upload_products():
    try:
        preview_data = session.get('upload_preview')
        if not preview_data:
            return jsonify({'success': False, 'error': 'Нет данных для загрузки'})
        
        products = load_products()
        existing_articles = {p['article']: p for p in products}
        
        new_count = 0
        updated_count = 0
        
        for item in preview_data['products']:
            article = item['article']
            
            if article in existing_articles:
                existing_articles[article]['name'] = item['name']
                if item.get('model'):
                    existing_articles[article]['model'] = item['model']
                updated_count += 1
            else:
                new_product = {
                    'name': item['name'],
                    'article': article
                }
                if item.get('model'):
                    new_product['model'] = item['model']
                products.append(new_product)
                new_count += 1
        
        if save_products(products):
            session.pop('upload_preview', None)
            
            stocks = load_stocks()
            stock_updated = 0
            for item in preview_data['products']:
                if item.get('stock') is not None:
                    article = item['article']
                    current = stocks.get(article, 0)
                    if item['stock'] != current:
                        stocks[article] = item['stock']
                        stock_updated += 1
                        
                        order_data = {
                            "Operation Type": "Инвентаризация",
                            "Date": datetime.now().strftime("%Y-%m-%d"),
                            "Order Number": "Загрузка из файла",
                            "Product Name": item['name'],
                            "Product Article": article,
                            "Quantity": item['stock']
                        }
                        save_order(order_data)
            
            if stock_updated > 0:
                save_stocks(stocks)
            
            metrics.record_operation('products_bulk_upload')
            return jsonify({
                'success': True,
                'new': new_count,
                'updated': updated_count,
                'stock_updated': stock_updated
            })
        else:
            return jsonify({'success': False, 'error': 'Ошибка при сохранении'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/upload/confirm-new', methods=['POST'])
@login_required
def confirm_upload_new_products():
    try:
        preview_data = session.get('new_products_preview')
        if not preview_data:
            return jsonify({'success': False, 'error': 'Нет данных для загрузки'})
        
        products = load_products()
        
        new_count = 0
        for item in preview_data['products']:
            new_product = {
                'name': item['name'],
                'article': item['article']
            }
            if item.get('model'):
                new_product['model'] = item['model']
            products.append(new_product)
            new_count += 1
        
        if save_products(products):
            session.pop('new_products_preview', None)
            
            stocks = load_stocks()
            stock_updated = 0
            for item in preview_data['products']:
                if item.get('stock') is not None:
                    article = item['article']
                    current = stocks.get(article, 0)
                    if item['stock'] != current:
                        stocks[article] = item['stock']
                        stock_updated += 1
                        
                        order_data = {
                            "Operation Type": "Инвентаризация",
                            "Date": datetime.now().strftime("%Y-%m-%d"),
                            "Order Number": "Загрузка из файла",
                            "Product Name": item['name'],
                            "Product Article": article,
                            "Quantity": item['stock']
                        }
                        save_order(order_data)
            
            if stock_updated > 0:
                save_stocks(stocks)
            
            metrics.record_operation('new_products_upload')
            return jsonify({
                'success': True,
                'new': new_count,
                'stock_updated': stock_updated
            })
        else:
            return jsonify({'success': False, 'error': 'Ошибка при сохранении'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/products/upload/template')
@login_required
def download_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Товары"
        
        headers = ["Название", "Артикул", "Модель", "Цена", "Остаток", "Категория", "Описание"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        examples = [
            ["Смартфон Xiaomi Redmi Note 10", "XIAOMI-RN10", "Redmi Note 10", 15000, 50, "Смартфоны", "6.4 дюйма, 128GB"],
            ["Ноутбук Lenovo IdeaPad 3", "LENOVO-IP3", "IdeaPad 3", 45000, 25, "Ноутбуки", "15.6 дюйма, 8GB RAM"],
            ["100мл банка 90250_40051", "100ml_jar_brown_St", "100ml_jar_brown_St", None, None, "Банки", ""],
        ]
        
        for row_num, example in enumerate(examples, 2):
            for col_num, value in enumerate(example, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        filename = f"template_products_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ ЗАКАЗОВ ====================

@app.route('/orders')
@login_required
def orders_list():
    try:
        orders = load_orders()
        
        grouped_orders = {}
        for order in orders:
            order_num = order.get('Order Number', 'Без номера')
            if order_num not in grouped_orders:
                grouped_orders[order_num] = {
                    'number': order_num,
                    'date': order.get('Date', ''),
                    'type': order.get('Operation Type', ''),
                    'items': [],
                    'total_items': 0,
                    'total_quantity': 0
                }
            
            grouped_orders[order_num]['items'].append(order)
            grouped_orders[order_num]['total_items'] += 1
            grouped_orders[order_num]['total_quantity'] += int(order.get('Quantity', 0))
        
        orders_list = []
        for order_num, order_data in grouped_orders.items():
            orders_list.append({
                'number': str(order_data.get('number', '')),
                'date': str(order_data.get('date', '')),
                'type': order_data.get('type', ''),
                'items': order_data.get('items', []),
                'total_items': order_data.get('total_items', 0),
                'total_quantity': order_data.get('total_quantity', 0)
            })
        
        orders_list.sort(key=lambda x: x['date'], reverse=True)
        
        today = datetime.now().date()
        thirty_days_ago = today - timedelta(days=30)
        
        return render_template('orders.html', 
                             orders=orders_list,
                             today=today.strftime('%Y-%m-%d'),
                             thirty_days_ago=thirty_days_ago.strftime('%Y-%m-%d'))
        
    except Exception as e:
        flash('Ошибка при загрузке списка заказов', 'danger')
        today = datetime.now().date().strftime('%Y-%m-%d')
        thirty_days_ago = (datetime.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')
        return render_template('orders.html', orders=[], today=today, thirty_days_ago=thirty_days_ago)


@app.route('/orders/filter', methods=['POST'])
@login_required
def filter_orders():
    try:
        data = request.json
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        order_type = data.get('type', 'all')
        search = data.get('search', '').lower()
        
        orders = load_orders()
        
        if start_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start = datetime.now().date() - timedelta(days=30)
        
        if end_date:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end = datetime.now().date()
        
        filtered_orders = []
        
        for order in orders:
            try:
                order_date = datetime.strptime(order['Date'], '%Y-%m-%d').date()
                if not (start <= order_date <= end):
                    continue
                
                if order_type != 'all' and order['Operation Type'] != order_type:
                    continue
                
                if search:
                    search_fields = [
                        str(order.get('Order Number', '')).lower(),
                        str(order.get('Product Name', '')).lower(),
                        str(order.get('Product Article', '')).lower()
                    ]
                    if not any(search in field for field in search_fields):
                        continue
                
                filtered_orders.append(order)
                
            except Exception as e:
                continue
        
        grouped_orders = {}
        for order in filtered_orders:
            order_num = order.get('Order Number', 'Без номера')
            if order_num not in grouped_orders:
                grouped_orders[order_num] = {
                    'number': order_num,
                    'date': order.get('Date', ''),
                    'type': order.get('Operation Type', ''),
                    'items': [],
                    'total_items': 0,
                    'total_quantity': 0
                }
            
            grouped_orders[order_num]['items'].append(order)
            grouped_orders[order_num]['total_items'] += 1
            grouped_orders[order_num]['total_quantity'] += int(order.get('Quantity', 0))
        
        orders_list = []
        for order_num, order_data in grouped_orders.items():
            orders_list.append({
                'number': order_data['number'],
                'date': order_data['date'],
                'type': order_data['type'],
                'items': order_data['items'],
                'total_items': order_data['total_items'],
                'total_quantity': order_data['total_quantity']
            })
        
        orders_list.sort(key=lambda x: x['date'], reverse=True)
        
        return jsonify({
            'success': True,
            'orders': orders_list,
            'total': len(orders_list),
            'total_quantity': sum(o['total_quantity'] for o in orders_list)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/order/<order_number>')
@login_required
def order_detail(order_number):
    try:
        orders = load_orders()
        order_items = [o for o in orders if o.get('Order Number') == order_number]
        
        if not order_items:
            flash('Заказ не найден', 'danger')
            return redirect(url_for('orders_list'))
        
        return render_template('order_detail.html', order_number=order_number, items=order_items)
        
    except Exception as e:
        flash('Ошибка при загрузке заказа', 'danger')
        return redirect(url_for('orders_list'))


@app.route('/order/<order_number>/update', methods=['POST'])
@login_required
def update_order(order_number):
    try:
        data = request.json
        items = data.get('items', [])
        
        orders = load_orders()
        
        orders = [o for o in orders if o.get('Order Number') != order_number]
        
        for item in items:
            order_data = {
                "Operation Type": "Расход",
                "Date": datetime.now().strftime("%Y-%m-%d"),
                "Order Number": order_number,
                "Product Name": item['name'],
                "Product Article": item['article'],
                "Quantity": item['quantity']
            }
            orders.append(order_data)
        
        if save_all_orders(orders):
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Ошибка при сохранении'})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ ИНВЕНТАРИЗАЦИИ ====================

@app.route('/inventory')
@login_required
def inventory():
    products = load_products()
    stocks = load_stocks()
    
    inventory_data = []
    for product in products:
        inventory_data.append({
            'name': product['name'],
            'article': product['article'],
            'model': product.get('model', ''),
            'current_stock': stocks.get(product['article'], 0)
        })
    
    return render_template('inventory.html', inventory=inventory_data)


@app.route('/inventory/update', methods=['POST'])
@login_required
def update_inventory():
    try:
        data = request.json
        stocks = load_stocks()
        
        for item in data['items']:
            article = item['article']
            new_stock = int(item['new_stock'])
            
            order_data = {
                "Operation Type": "Инвентаризация",
                "Date": datetime.now().strftime("%Y-%m-%d"),
                "Order Number": "Инвентаризация",
                "Product Name": item['name'],
                "Product Article": article,
                "Quantity": new_stock
            }
            save_order(order_data)
            
            stocks[article] = new_stock
        
        save_stocks(stocks)
        metrics.record_operation('inventory_update')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/inventory/preview', methods=['POST'])
@login_required
def preview_inventory():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        if file_ext not in {'.xlsx', '.xls', '.csv'}:
            return jsonify({'success': False, 'error': 'Неверный формат файла. Разрешены: .xlsx, .xls, .csv'})
        
        try:
            if file_ext == '.csv':
                encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'windows-1251']
                df = None
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file, encoding=encoding)
                        break
                    except:
                        continue
                if df is None:
                    encoding = detect_encoding(file)
                    df = pd.read_csv(file, encoding=encoding)
            else:
                df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'success': False, 'error': f'Ошибка при чтении файла: {str(e)}'})
        
        if df.empty:
            return jsonify({'success': False, 'error': 'Файл не содержит данных'})
        
        df = clean_dataframe(df)
        
        col_mapping = {
            'article': None,
            'name': None,
            'actual': None
        }
        
        for i, col in enumerate(df.columns):
            col_lower = str(col).lower().strip()
            
            if any(x in col_lower for x in ['артикул', 'article', 'код', 'sku']):
                col_mapping['article'] = i
            elif any(x in col_lower for x in ['название', 'товар', 'name', 'product']):
                col_mapping['name'] = i
            elif any(x in col_lower for x in ['фактический', 'actual', 'новый', 'остаток', 'stock']):
                col_mapping['actual'] = i
        
        if col_mapping['article'] is None:
            return jsonify({'success': False, 'error': 'Не найдена колонка "Артикул"'})
        
        if col_mapping['actual'] is None:
            return jsonify({'success': False, 'error': 'Не найдена колонка "Фактический остаток"'})
        
        stocks = load_stocks()
        products = load_products()
        
        preview_items = []
        errors = []
        
        for idx, row in df.iterrows():
            try:
                article_col = df.columns[col_mapping['article']]
                article = str(row[article_col]).strip()
                
                if pd.isna(article) or not article or article == 'nan':
                    continue
                
                actual_col = df.columns[col_mapping['actual']]
                actual_value = row[actual_col]
                
                if pd.isna(actual_value):
                    actual = 0
                else:
                    try:
                        if isinstance(actual_value, (int, float)):
                            actual = int(actual_value)
                        else:
                            clean_value = re.sub(r'[^\d\-]', '', str(actual_value))
                            actual = int(clean_value) if clean_value else 0
                    except:
                        actual = 0
                
                name = ''
                if col_mapping['name'] is not None:
                    name_col = df.columns[col_mapping['name']]
                    if not pd.isna(row[name_col]):
                        name = str(row[name_col]).strip()
                
                if not name:
                    for product in products:
                        if product['article'] == article:
                            name = product['name']
                            break
                
                current = stocks.get(article, 0)
                
                preview_items.append({
                    'article': article,
                    'name': name,
                    'current_stock': current,
                    'new_stock': actual
                })
                
            except Exception as e:
                errors.append(f"Строка {idx + 2}: {str(e)}")
        
        return jsonify({
            'success': True,
            'items': preview_items,
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/inventory/upload', methods=['POST'])
@login_required
def upload_inventory():
    try:
        data = request.json
        items = data.get('items', [])
        
        if not items:
            return jsonify({'success': False, 'error': 'Нет данных для обновления'})
        
        stocks = load_stocks()
        products = load_products()
        
        updated_count = 0
        new_products_count = 0
        errors = []
        
        for item in items:
            try:
                article = item['article']
                new_stock = int(item['new_stock'])
                name = item.get('name', '').strip()
                
                product_exists = False
                existing_name = ''
                
                for product in products:
                    if product['article'] == article:
                        product_exists = True
                        existing_name = product['name']
                        break
                
                if not product_exists:
                    if not name:
                        name = f"Товар {article}"
                    
                    products.append({'name': name, 'article': article})
                    new_products_count += 1
                else:
                    name = existing_name
                
                order_data = {
                    "Operation Type": "Инвентаризация",
                    "Date": datetime.now().strftime("%Y-%m-%d"),
                    "Order Number": "Инвентаризация из файла",
                    "Product Name": name,
                    "Product Article": article,
                    "Quantity": new_stock
                }
                
                save_order(order_data)
                
                stocks[article] = new_stock
                updated_count += 1
                
            except Exception as e:
                errors.append(f"Товар {item.get('article', 'неизвестно')}: {str(e)}")
        
        if new_products_count > 0:
            save_products(products)
        
        save_stocks(stocks)
        
        metrics.record_operation('inventory_upload')
        return jsonify({
            'success': True,
            'updated': updated_count,
            'new_products': new_products_count,
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ ОТЧЕТОВ ====================

@app.route('/reports')
@login_required
def reports():
    today = datetime.now().date()
    first_day_of_month = today.replace(day=1)
    
    return render_template('reports.html',
                         today=today.strftime('%Y-%m-%d'),
                         first_day_of_month=first_day_of_month.strftime('%Y-%m-%d'))


@app.route('/reports/composite')
@login_required
def composite_report_page():
    """Страница отчета по составным артикулам"""
    return render_template('composite_report.html')


@app.route('/reports/generate', methods=['POST'])
@login_required
def generate_report():
    try:
        data = request.json
        start_date = datetime.strptime(data['start_date'], "%Y-%m-%d").date()
        end_date = datetime.strptime(data['end_date'], "%Y-%m-%d").date()
        report_type = data.get('report_type', 'all')
        article_filter = data.get('article', '')
        
        orders = load_orders()
        
        filtered_orders = []
        for order in orders:
            try:
                order_date = datetime.strptime(order['Date'], "%Y-%m-%d").date()
                if start_date <= order_date <= end_date:
                    if report_type == 'all' or order['Operation Type'] == report_type:
                        if not article_filter or order['Product Article'] == article_filter:
                            filtered_orders.append(order)
            except Exception as e:
                continue
        
        summary = {}
        detailed = []
        composite_stats = {'total': 0, 'orders': []}
        
        for order in filtered_orders:
            article = order['Product Article']
            quantity = int(order['Quantity'])
            
            detailed.append({
                'date': order['Date'],
                'order_number': order['Order Number'],
                'type': order['Operation Type'],
                'product_name': order['Product Name'],
                'article': article,
                'quantity': quantity,
                'is_composite': '_' in article if article else False
            })
            
            if '_' in article:
                composite_stats['total'] += 1
                if order['Order Number'] not in composite_stats['orders']:
                    composite_stats['orders'].append(order['Order Number'])
            
            if article not in summary:
                summary[article] = {
                    'article': article,
                    'name': order['Product Name'],
                    'total_in': 0,
                    'total_out': 0,
                    'inventory': None,
                    'operations': [],
                    'is_composite': '_' in article
                }
            
            summary[article]['operations'].append({
                'date': order['Date'],
                'type': order['Operation Type'],
                'quantity': quantity
            })
            
            if order['Operation Type'] == 'Приход':
                summary[article]['total_in'] += quantity
            elif order['Operation Type'] == 'Расход':
                summary[article]['total_out'] += quantity
            elif order['Operation Type'] == 'Инвентаризация':
                summary[article]['inventory'] = quantity
        
        summary_list = []
        for article, data in summary.items():
            if data['inventory'] is not None:
                final_stock = data['inventory']
            else:
                final_stock = data['total_in'] - data['total_out']
            
            summary_list.append({
                'article': article,
                'name': data['name'],
                'total_in': data['total_in'],
                'total_out': data['total_out'],
                'inventory': data['inventory'],
                'final_stock': final_stock,
                'operations': data['operations'],
                'is_composite': data['is_composite']
            })
        
        return jsonify({
            'success': True, 
            'summary': summary_list,
            'detailed': detailed,
            'composite_stats': composite_stats
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/reports/composite', methods=['POST'])
@login_required
def composite_report():
    """Генерация отчета по составным артикулам с фильтрацией по части"""
    try:
        data = request.json
        part = data.get('part', '').strip()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not part:
            return jsonify({'success': False, 'error': 'Укажите часть артикула для поиска'})
        
        # Загружаем данные
        products = load_products()
        stocks = load_stocks()
        orders = load_orders()
        
        # Находим все составные артикулы, содержащие искомую часть
        matching_articles = []
        article_to_product = {}
        
        for product in products:
            article = product['article']
            # Проверяем, содержит ли артикул искомую часть
            if part.lower() in article.lower():
                matching_articles.append(article)
                article_to_product[article] = product
        
        app.logger.info(f"Найдено артикулов с '{part}': {len(matching_articles)}")
        
        # Фильтруем заказы по дате
        filtered_orders = []
        for order in orders:
            try:
                order_date = datetime.strptime(order['Date'], "%Y-%m-%d").date()
                
                if start_date:
                    start = datetime.strptime(start_date, '%Y-%m-%d').date()
                    if order_date < start:
                        continue
                
                if end_date:
                    end = datetime.strptime(end_date, '%Y-%m-%d').date()
                    if order_date > end:
                        continue
                
                filtered_orders.append(order)
            except:
                continue
        
        app.logger.info(f"Заказов за период: {len(filtered_orders)}")
        
        # Собираем статистику по каждому артикулу
        composite_stats = []
        total_in = 0
        total_out = 0
        
        for article in matching_articles:
            product = article_to_product[article]
            
            # Считаем операции по этому артикулу
            article_in = 0
            article_out = 0
            operations = []
            
            for order in filtered_orders:
                if order.get('Product Article') == article:
                    quantity = int(order.get('Quantity', 0))
                    op_type = order.get('Operation Type', '')
                    
                    operations.append({
                        'date': order.get('Date', ''),
                        'type': op_type,
                        'quantity': quantity,
                        'order_number': order.get('Order Number', '')
                    })
                    
                    if op_type == 'Приход':
                        article_in += quantity
                        total_in += quantity
                    elif op_type == 'Расход':
                        article_out += quantity
                        total_out += quantity
            
            # Разбираем составной артикул на части
            parts = article.split('_') if '_' in article else [article]
            
            # Находим позицию искомой части
            part_positions = []
            for i, part_item in enumerate(parts):
                if part.lower() in part_item.lower():
                    part_positions.append({
                        'index': i,
                        'value': part_item,
                        'full_match': part_item.lower() == part.lower()
                    })
            
            composite_stats.append({
                'article': article,
                'name': product['name'],
                'model': product.get('model', ''),
                'current_stock': stocks.get(article, 0),
                'total_in': article_in,
                'total_out': article_out,
                'operations': operations,
                'operations_count': len(operations),
                'parts': parts,
                'part_positions': part_positions,
                'has_matching_part': len(part_positions) > 0
            })
        
        # Сортируем: сначала те, где есть точное совпадение части
        composite_stats.sort(key=lambda x: (
            - (1 if any(p.get('full_match') for p in x['part_positions']) else 0),
            -x['total_out'],
            x['article']
        ))
        
        # Собираем уникальные части артикулов для статистики
        all_parts = {}
        for stat in composite_stats:
            for i, part_item in enumerate(stat['parts']):
                if part_item not in all_parts:
                    all_parts[part_item] = {
                        'part': part_item,
                        'count': 0,
                        'articles': [],
                        'total_stock': 0,
                        'total_out': 0
                    }
                all_parts[part_item]['count'] += 1
                all_parts[part_item]['articles'].append(stat['article'])
                all_parts[part_item]['total_stock'] += stat['current_stock']
                all_parts[part_item]['total_out'] += stat['total_out']
        
        # Сортируем части по популярности
        sorted_parts = sorted(
            [p for p in all_parts.values() if p['count'] > 1],
            key=lambda x: (-x['count'], -x['total_out'])
        )
        
        metrics.record_operation('composite_report')
        return jsonify({
            'success': True,
            'part': part,
            'total_articles': len(composite_stats),
            'stats': composite_stats,
            'parts_summary': sorted_parts[:50],  # Топ-50 частей
            'summary': {
                'total_in': total_in,
                'total_out': total_out,
                'unique_articles': len(composite_stats),
                'unique_parts': len(sorted_parts)
            }
        })
        
    except Exception as e:
        app.logger.error(f"Ошибка при генерации отчета по составным артикулам: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ PDF ====================

@app.route('/upload/pdf', methods=['GET', 'POST'])
@login_required
def upload_pdf():
    if request.method == 'GET':
        return render_template('upload_pdf.html')
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Неверный формат файла. Разрешены только PDF'})
    
    filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    try:
        orders = parse_pdf_orders(filepath)
        
        if os.path.exists(filepath):
            os.remove(filepath)
        
        if not orders:
            return jsonify({'success': False, 'error': 'Не удалось найти заказы в PDF файле'})
        
        return jsonify({
            'success': True,
            'orders': orders
        })
        
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'success': False, 'error': str(e)})


@app.route('/upload/pdf/load', methods=['POST'])
@login_required
def load_pdf_orders():
    try:
        data = request.json
        orders = data.get('orders', [])
        
        if not orders:
            return jsonify({'success': False, 'error': 'Нет заказов для загрузки'})
        
        app.logger.info("=" * 60)
        app.logger.info("ЗАПУСК ЗАГРУЗКИ ЗАКАЗОВ ИЗ PDF")
        app.logger.info(f"Получено заказов: {len(orders)}")
        app.logger.info("=" * 60)
        
        products = load_products()
        stocks = load_stocks()
        processed_orders = load_processed_orders()
        
        loaded_count = 0
        skipped_count = 0
        new_products_count = 0
        composite_count = 0
        missing_products = []
        composite_log = []
        
        for order in orders:
            order_number = str(order.get('order_number', ''))
            if not order_number:
                continue
                
            if order_number in processed_orders:
                skipped_count += 1
                app.logger.info(f"Заказ {order_number} уже обработан, пропускаем")
                continue
            
            items = order.get('items', [])
            if not items:
                continue
                
            app.logger.info(f"Обработка заказа {order_number}, товаров: {len(items)}")
            
            for item in items:
                composite_article = str(item.get('article', '') or item.get('Артикул', '')).strip()
                model = str(item.get('model', '') or item.get('Модель', '')).strip()
                product_name = str(item.get('product', '') or item.get('Товар', '') or item.get('product_name', '')).strip()
                quantity = int(item.get('quantity', 0))
                
                app.logger.info(f"  Товар: арт='{composite_article}', модель='{model}', название='{product_name}', кол-во={quantity}")
                
                if not composite_article and not model and not product_name:
                    app.logger.warning("    Пропуск: нет идентификаторов товара")
                    continue
                
                if quantity <= 0:
                    app.logger.warning(f"    Пропуск: некорректное количество {quantity}")
                    continue
                
                # Создаем объект для счетчика новых товаров
                new_count_ref = {'count': 0}
                
                if composite_article and '_' in composite_article:
                    composite_count += 1
                    
                    log_entry = {
                        'original': composite_article,
                        'order': order_number,
                        'quantity': quantity
                    }
                    
                    app.logger.info(f"    Обработка составного артикула: {composite_article}")
                    
                    found_count, created_count = process_composite_article(
                        composite_article, products, stocks, order_number, 
                        quantity, product_name, model,
                        new_products_count_ref=new_count_ref,
                        missing_products=missing_products
                    )
                    
                    new_products_count += new_count_ref['count']
                    
                    log_entry['found'] = found_count
                    log_entry['created'] = created_count
                    composite_log.append(log_entry)
                    
                    app.logger.info(f"    Результат: найдено={found_count}, создано={created_count}")
                    
                else:
                    app.logger.info(f"    Обработка одиночного артикула: {composite_article or model or product_name}")
                    
                    success = process_single_article(
                        composite_article, model, product_name, quantity,
                        products, stocks, order_number, 
                        display_name=product_name,
                        new_products_count_ref=new_count_ref,
                        missing_products=missing_products
                    )
                    
                    new_products_count += new_count_ref['count']
                    
                    app.logger.info(f"    Результат: {'успешно' if success else 'ошибка'}")
            
            # Помечаем заказ как обработанный
            save_processed_order(order_number)
            loaded_count += 1
            app.logger.info(f"Заказ {order_number} успешно обработан")
        
        # Сохраняем изменения
        if new_products_count > 0:
            app.logger.info(f"Сохраняем {new_products_count} новых товаров")
            save_products(products)
        
        app.logger.info("Сохраняем остатки")
        save_stocks(stocks)
        
        # Сохраняем лог составных артикулов
        if composite_log:
            log_file = os.path.join(app.config['DATA_FOLDER'], 'composite_articles_log.json')
            try:
                with open(log_file, 'w', encoding='utf-8') as f:
                    json.dump(composite_log, f, indent=2, ensure_ascii=False)
                app.logger.info(f"Лог составных артикулов сохранен в {log_file}")
            except Exception as e:
                app.logger.error(f"Ошибка сохранения лога: {e}")
        
        app.logger.info("=" * 60)
        app.logger.info(f"ИТОГИ ЗАГРУЗКИ: загружено={loaded_count}, пропущено={skipped_count}, новых товаров={new_products_count}")
        app.logger.info("=" * 60)
        
        metrics.record_operation('pdf_orders_loaded')
        return jsonify({
            'success': True,
            'loaded': loaded_count,
            'skipped': skipped_count,
            'new_products': new_products_count,
            'composite_articles': composite_count,
            'composite_log': composite_log[:20],
            'missing_products': missing_products[:10]
        })
        
    except Exception as e:
        app.logger.error(f"Ошибка при загрузке заказов из PDF: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


def detect_table_structure(table):
    """Определяет структуру таблицы по заголовкам"""
    structure = {
        'product_col': None,
        'article_col': None,
        'quantity_col': None,
        'model_col': None
    }
    
    if not table or len(table) < 1:
        return structure
    
    headers = table[0]
    for i, header in enumerate(headers):
        if not header:
            continue
        header_lower = str(header).lower()
        
        if any(x in header_lower for x in ['товар', 'наименование', 'product']):
            structure['product_col'] = i
        elif any(x in header_lower for x in ['артикул', 'article', 'код']):
            structure['article_col'] = i
        elif any(x in header_lower for x in ['количество', 'кол-во', 'quantity']):
            structure['quantity_col'] = i
        elif any(x in header_lower for x in ['модель', 'model']):
            structure['model_col'] = i
    
    return structure


def extract_order_number(text):
    """Извлекает номер заказа из текста"""
    patterns = [
        r'№\s*заказа:?\s*(\d+)',
        r'Заказ\s*№?\s*(\d+)',
        r'Order\s*#?\s*(\d+)',
        r'Номер\s*заказа:?\s*(\d+)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def parse_pdf_orders(pdf_path):
    """Улучшенный парсер PDF с поддержкой различных форматов"""
    orders = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                # Извлекаем текст и таблицы
                text = page.extract_text() or ""
                tables = page.extract_tables()
                
                # Ищем номер заказа в тексте страницы
                order_number = extract_order_number(text)
                
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    
                    # Определяем структуру таблицы
                    table_structure = detect_table_structure(table)
                    
                    # Извлекаем товары
                    items = extract_items_from_table(table, table_structure)
                    
                    if items:
                        orders.append({
                            'order_number': order_number or f"PDF-{len(orders)+1}",
                            'items': items
                        })
        
        return orders
    except Exception as e:
        app.logger.error(f"Ошибка парсинга PDF: {e}")
        return []


def extract_items_from_table(table, structure):
    """Извлекает товары из таблицы PDF"""
    items = []
    
    for row in table[1:]:  # Пропускаем заголовок
        if not row:
            continue
        
        item = {'quantity': 1}
        
        # Извлекаем название товара
        if structure['product_col'] is not None and structure['product_col'] < len(row):
            product = row[structure['product_col']]
            if product and str(product).strip():
                item['product'] = str(product).strip()
        
        # Извлекаем артикул
        if structure['article_col'] is not None and structure['article_col'] < len(row):
            article = row[structure['article_col']]
            if article and str(article).strip():
                item['article'] = str(article).strip()
        
        # Извлекаем модель
        if structure['model_col'] is not None and structure['model_col'] < len(row):
            model = row[structure['model_col']]
            if model and str(model).strip():
                item['model'] = str(model).strip()
        
        # Извлекаем количество
        if structure['quantity_col'] is not None and structure['quantity_col'] < len(row):
            quantity = row[structure['quantity_col']]
            if quantity:
                # Извлекаем числа из строки
                numbers = re.findall(r'\d+', str(quantity))
                if numbers:
                    item['quantity'] = int(numbers[0])
        
        # Проверяем, что у нас есть достаточно данных
        if ('article' in item or 'product' in item or 'model' in item) and item['quantity'] > 0:
            items.append(item)
    
    return items


@app.route('/debug/pdf-parser', methods=['POST'])
@login_required
def debug_pdf_parser():
    """Диагностика парсера PDF"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'Файл не выбран'})
        
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'error': 'Неверный формат файла'})
        
        filename = secure_filename(f"debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        debug_info = {
            'file': file.filename,
            'size': os.path.getsize(filepath),
            'pages': [],
            'extracted_text': [],
            'tables': []
        }
        
        try:
            with pdfplumber.open(filepath) as pdf:
                debug_info['pages_count'] = len(pdf.pages)
                
                for page_num, page in enumerate(pdf.pages):
                    page_info = {
                        'number': page_num + 1,
                        'text': page.extract_text() or '',
                        'tables_count': len(page.extract_tables())
                    }
                    debug_info['pages'].append(page_info)
                    
                    # Сохраняем первые 500 символов текста
                    if page_info['text']:
                        debug_info['extracted_text'].append(page_info['text'][:500])
                    
                    # Извлекаем таблицы
                    tables = page.extract_tables()
                    for table_idx, table in enumerate(tables):
                        if table and len(table) > 1:
                            table_info = {
                                'page': page_num + 1,
                                'index': table_idx,
                                'rows': len(table),
                                'cols': len(table[0]) if table[0] else 0,
                                'sample': table[:3]  # Первые 3 строки
                            }
                            debug_info['tables'].append(table_info)
            
            # Также пробуем распарсить заказы
            orders = parse_pdf_orders(filepath)
            debug_info['parsed_orders'] = len(orders)
            debug_info['orders_sample'] = orders[:2] if orders else []
            
        except Exception as e:
            debug_info['parse_error'] = str(e)
        
        # Удаляем файл
        if os.path.exists(filepath):
            os.remove(filepath)
        
        return jsonify({
            'success': True,
            'debug_info': debug_info
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ OZON ====================

@app.route('/upload/ozon', methods=['GET', 'POST'])
@login_required
def upload_ozon():
    """Загрузка Excel файла с заказами Ozon"""
    if request.method == 'GET':
        return render_template('upload_ozon.html')
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in {'.xlsx', '.xls'}:
        return jsonify({'success': False, 'error': 'Неверный формат файла. Разрешены только Excel файлы (.xlsx, .xls)'})
    
    filename = secure_filename(f"ozon_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    try:
        orders = parse_ozon_excel(filepath)
        
        if os.path.exists(filepath):
            os.remove(filepath)
        
        if not orders:
            return jsonify({
                'success': False,
                'error': 'Не удалось найти заказы в Excel файле. Проверьте структуру файла.'
            })
        
        return jsonify({
            'success': True,
            'orders': orders,
            'total_orders': len(orders),
            'total_items': sum(len(o['items']) for o in orders)
        })
        
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        app.logger.error(f"Ошибка при обработке Excel: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


def parse_ozon_excel(filepath):
    """
    Парсит Excel файл с заказами Ozon с правильным расчетом количества
    Извлекает количество из названия товара (например, "10 штук" -> 10)
    и умножает на количество в колонке "Количество"
    """
    orders = []
    
    try:
        df = pd.read_excel(filepath)
        app.logger.info(f"Excel файл прочитан. Колонки: {list(df.columns)}")
        app.logger.info(f"Всего строк: {len(df)}")
        
        # Приводим названия колонок к нижнему регистру для поиска
        df.columns = [str(col).strip().lower() for col in df.columns]
        
        # Определяем нужные колонки
        article_col = None
        name_col = None
        quantity_col = None
        order_num_col = None
        date_col = None
        product_name_col = None  # Колонка с названием товара (Наименование товара)
        
        for col in df.columns:
            col_lower = col.lower()
            
            if any(x in col_lower for x in ['артикул', 'article', 'sku', 'код']):
                article_col = col
                app.logger.info(f"Найдена колонка артикула: {col}")
            
            elif any(x in col_lower for x in ['наименование', 'товар', 'name', 'product']):
                if 'наименование товара' in col_lower or 'наименование' in col_lower:
                    product_name_col = col  # Это колонка с полным названием товара
                    app.logger.info(f"Найдена колонка наименования товара: {col}")
                else:
                    name_col = col
                    app.logger.info(f"Найдена колонка названия: {col}")
            
            elif any(x in col_lower for x in ['количество', 'кол-во', 'quantity', 'qty']):
                quantity_col = col
                app.logger.info(f"Найдена колонка количества: {col}")
            
            elif any(x in col_lower for x in ['номер заказа', 'заказ', 'order', 'order number', 'номер отправления']):
                order_num_col = col
                app.logger.info(f"Найдена колонка номера заказа: {col}")
            
            elif any(x in col_lower for x in ['дата', 'date']):
                date_col = col
                app.logger.info(f"Найдена колонка даты: {col}")
        
        # Если не нашли специальную колонку "Наименование товара", используем name_col
        if not product_name_col and name_col:
            product_name_col = name_col
        
        if not product_name_col and len(df.columns) >= 2:
            product_name_col = df.columns[1]
            app.logger.info(f"Колонка наименования товара не найдена, используем вторую: {product_name_col}")
        
        if not article_col and len(df.columns) >= 1:
            article_col = df.columns[0]
            app.logger.info(f"Колонка артикула не найдена, используем первую: {article_col}")
        
        if not quantity_col and len(df.columns) >= 3:
            quantity_col = df.columns[2]
            app.logger.info(f"Колонка количества не найдена, используем третью: {quantity_col}")
        
        # Функция для извлечения количества из названия товара
        def extract_quantity_from_name(product_name):
            """
            Извлекает количество из названия товара
            Примеры:
            "Флакон 100 мл стеклянный с пластиковой крышкой, 10 штук" -> 10
            "Баночки стеклянные с крышкой, 15 мл 3 штуки" -> 3
            "Флакон с капельницей 2,5 мл стеклянный, 10 штук." -> 10
            "Флакон для диффузора, стеклянный, 30 мл, пустой, 3 шт" -> 3
            "Флакон 5 мл стеклянный с крышкой с пипеткой, 10 штук." -> 10
            """
            if not product_name or pd.isna(product_name):
                return 1
            
            product_name = str(product_name).lower()
            app.logger.debug(f"Извлечение количества из: {product_name}")
            
            # Паттерны для поиска количества
            patterns = [
                r'(\d+)\s*шт\.?ук?',  # 10 штук, 10 шт, 10шт
                r'(\d+)\s*штук',       # 10 штук
                r'(\d+)\s*шт',          # 10 шт
                r',\s*(\d+)\s*шт',      # , 10 шт
                r'(\d+)\s*набор',       # 3 набора
                r'(\d+)\s*единиц',      # 5 единиц
                r'(\d+)\s*компл',       # 2 комплекта
                r'в количестве\s*(\d+)', # в количестве 10
                r'кол-во:\s*(\d+)',      # кол-во: 10
                r'(\d+)\s*$'             # число в конце строки
            ]
            
            for pattern in patterns:
                match = re.search(pattern, product_name)
                if match:
                    try:
                        quantity = int(match.group(1))
                        app.logger.debug(f"  Найдено количество: {quantity} (паттерн: {pattern})")
                        return quantity
                    except:
                        continue
            
            # Если не нашли по паттернам, ищем все числа в строке
            numbers = re.findall(r'\d+', product_name)
            if numbers:
                # Берем последнее число (часто это количество в упаковке)
                try:
                    quantity = int(numbers[-1])
                    app.logger.debug(f"  Найдено количество (последнее число): {quantity}")
                    return quantity
                except:
                    pass
            
            app.logger.debug("  Количество не найдено, используем 1")
            return 1
        
        # Функция для расчета итогового количества
        def calculate_total_quantity(row_quantity, pack_quantity):
            """
            Рассчитывает итоговое количество:
            количество из колонки * количество в упаковке из названия
            """
            try:
                # Количество из колонки (сколько упаковок)
                if pd.isna(row_quantity):
                    col_qty = 1
                else:
                    if isinstance(row_quantity, (int, float)):
                        col_qty = int(row_quantity)
                    else:
                        # Извлекаем число из строки
                        numbers = re.findall(r'\d+', str(row_quantity))
                        col_qty = int(numbers[0]) if numbers else 1
                
                # Итоговое количество
                total = col_qty * pack_quantity
                app.logger.debug(f"  Расчет: {col_qty} (упаковок) * {pack_quantity} (в упаковке) = {total}")
                return total
                
            except Exception as e:
                app.logger.error(f"Ошибка расчета количества: {e}")
                return 1
        
        # Группируем по заказам, если есть колонка с номером заказа
        if order_num_col:
            grouped = df.groupby(order_num_col)
            
            for order_num, group in grouped:
                order_items = []
                order_number = str(order_num)
                
                app.logger.info(f"Обработка заказа {order_number}, позиций: {len(group)}")
                
                for _, row in group.iterrows():
                    try:
                        # Получаем артикул
                        article = str(row[article_col]).strip() if article_col and not pd.isna(row[article_col]) else ''
                        
                        # Получаем полное название товара
                        product_full_name = ''
                        if product_name_col and not pd.isna(row[product_name_col]):
                            product_full_name = str(row[product_name_col]).strip()
                        
                        # Получаем название товара для отображения (краткое)
                        product_display = product_full_name
                        if len(product_display) > 50:
                            product_display = product_display[:50] + "..."
                        
                        # Получаем количество из колонки (сколько упаковок)
                        col_quantity = 1
                        if quantity_col and not pd.isna(row[quantity_col]):
                            try:
                                if isinstance(row[quantity_col], (int, float)):
                                    col_quantity = int(row[quantity_col])
                                else:
                                    numbers = re.findall(r'\d+', str(row[quantity_col]))
                                    col_quantity = int(numbers[0]) if numbers else 1
                            except:
                                col_quantity = 1
                        
                        # Извлекаем количество в упаковке из названия
                        pack_quantity = extract_quantity_from_name(product_full_name)
                        
                        # Рассчитываем итоговое количество
                        total_quantity = calculate_total_quantity(col_quantity, pack_quantity)
                        
                        # Получаем дату
                        date = ''
                        if date_col and not pd.isna(row[date_col]):
                            try:
                                if isinstance(row[date_col], datetime):
                                    date = row[date_col].strftime('%Y-%m-%d')
                                else:
                                    date = str(row[date_col]).strip()
                            except:
                                date = ''
                        
                        if article and total_quantity > 0:
                            item = {
                                'article': article,
                                'product': product_display,
                                'product_full_name': product_full_name,
                                'quantity': total_quantity,
                                'pack_quantity': pack_quantity,
                                'col_quantity': col_quantity,
                                'calculation': f"{col_quantity} упак × {pack_quantity} шт = {total_quantity}"
                            }
                            
                            if '_' in article:
                                app.logger.info(f"  Обнаружен составной артикул: {article}")
                            
                            app.logger.info(f"  Товар: {article}, расчет: {item['calculation']}")
                            order_items.append(item)
                            
                    except Exception as e:
                        app.logger.error(f"Ошибка обработки строки: {e}")
                        continue
                
                if order_items:
                    orders.append({
                        'order_number': order_number,
                        'date': date if date else datetime.now().strftime('%Y-%m-%d'),
                        'items': order_items,
                        'total_items': len(order_items),
                        'total_quantity': sum(i['quantity'] for i in order_items)
                    })
        else:
            # Если нет номера заказа, создаем один общий заказ
            order_items = []
            
            for idx, row in df.iterrows():
                try:
                    article = str(row[article_col]).strip() if article_col and not pd.isna(row[article_col]) else ''
                    
                    product_full_name = ''
                    if product_name_col and not pd.isna(row[product_name_col]):
                        product_full_name = str(row[product_name_col]).strip()
                    
                    product_display = product_full_name
                    if len(product_display) > 50:
                        product_display = product_display[:50] + "..."
                    
                    col_quantity = 1
                    if quantity_col and not pd.isna(row[quantity_col]):
                        try:
                            if isinstance(row[quantity_col], (int, float)):
                                col_quantity = int(row[quantity_col])
                            else:
                                numbers = re.findall(r'\d+', str(row[quantity_col]))
                                col_quantity = int(numbers[0]) if numbers else 1
                        except:
                            col_quantity = 1
                    
                    pack_quantity = extract_quantity_from_name(product_full_name)
                    total_quantity = calculate_total_quantity(col_quantity, pack_quantity)
                    
                    if article and total_quantity > 0:
                        item = {
                            'article': article,
                            'product': product_display,
                            'product_full_name': product_full_name,
                            'quantity': total_quantity,
                            'pack_quantity': pack_quantity,
                            'col_quantity': col_quantity,
                            'calculation': f"{col_quantity} упак × {pack_quantity} шт = {total_quantity}"
                        }
                        
                        if '_' in article:
                            app.logger.info(f"  Обнаружен составной артикул: {article}")
                        
                        order_items.append(item)
                        
                except Exception as e:
                    app.logger.error(f"Ошибка обработки строки {idx}: {e}")
                    continue
            
            if order_items:
                orders.append({
                    'order_number': f"OZON-{datetime.now().strftime('%Y%m%d%H%M%S')}",
                    'date': datetime.now().strftime('%Y-%m-%d'),
                    'items': order_items,
                    'total_items': len(order_items),
                    'total_quantity': sum(i['quantity'] for i in order_items)
                })
        
        # Логируем итоги
        app.logger.info(f"Всего найдено заказов Ozon: {len(orders)}")
        for i, order in enumerate(orders):
            app.logger.info(f"Заказ {i+1}: №{order['order_number']}, позиций: {order['total_items']}, всего единиц: {order['total_quantity']}")
            for item in order['items'][:3]:  # Показываем первые 3 товара для примера
                app.logger.info(f"  - {item['article']}: {item['calculation']}")
        
        return orders
        
    except Exception as e:
        app.logger.error(f"Ошибка при парсинге Excel: {e}")
        traceback.print_exc()
        return []


@app.route('/upload/ozon/load', methods=['POST'])
@login_required
def load_ozon_orders():
    """Загрузка выбранных заказов Ozon в систему"""
    try:
        data = request.json
        orders = data.get('orders', [])
        
        if not orders:
            return jsonify({'success': False, 'error': 'Нет заказов для загрузки'})
        
        app.logger.info("=" * 60)
        app.logger.info("ЗАПУСК ЗАГРУЗКИ ЗАКАЗОВ OZON")
        app.logger.info(f"Получено заказов: {len(orders)}")
        app.logger.info("=" * 60)
        
        products = load_products()
        stocks = load_stocks()
        processed_orders = load_processed_orders()
        
        loaded_count = 0
        skipped_count = 0
        new_products_count = 0
        composite_count = 0
        missing_products = []
        composite_log = []
        
        for order in orders:
            order_number = str(order.get('order_number', ''))
            if not order_number:
                continue
                
            if order_number in processed_orders:
                skipped_count += 1
                app.logger.info(f"Заказ {order_number} уже обработан, пропускаем")
                continue
            
            app.logger.info(f"Загрузка заказа Ozon №{order_number}")
            
            items = order.get('items', [])
            if not items:
                continue
                
            app.logger.info(f"  Позиций в заказе: {len(items)}")
            
            for item in items:
                composite_article = str(item.get('article', '')).strip()
                product_name = str(item.get('product', '')).strip()
                quantity = int(item.get('quantity', 0))
                calculation = item.get('calculation', '')
                
                app.logger.info(f"  Товар: арт='{composite_article}', название='{product_name}', кол-во={quantity} ({calculation})")
                
                if not composite_article:
                    app.logger.warning("    Пропуск: нет артикула")
                    continue
                
                if quantity <= 0:
                    app.logger.warning(f"    Пропуск: некорректное количество {quantity}")
                    continue
                
                new_count_ref = {'count': 0}
                
                if '_' in composite_article:
                    composite_count += 1
                    
                    log_entry = {
                        'original': composite_article,
                        'order': order_number,
                        'quantity': quantity,
                        'calculation': calculation
                    }
                    
                    app.logger.info(f"    Обработка составного артикула: {composite_article}")
                    
                    found_count, created_count = process_composite_article(
                        composite_article, products, stocks, order_number, 
                        quantity, product_name, '',
                        new_products_count_ref=new_count_ref,
                        missing_products=missing_products
                    )
                    
                    new_products_count += new_count_ref['count']
                    
                    log_entry['found'] = found_count
                    log_entry['created'] = created_count
                    composite_log.append(log_entry)
                    
                    app.logger.info(f"    Результат: найдено={found_count}, создано={created_count}")
                    
                else:
                    app.logger.info(f"    Обработка одиночного артикула: {composite_article}")
                    
                    process_single_article(
                        composite_article, '', product_name, quantity,
                        products, stocks, order_number, 
                        display_name=product_name,
                        new_products_count_ref=new_count_ref,
                        missing_products=missing_products
                    )
                    
                    new_products_count += new_count_ref['count']
                    app.logger.info(f"    Результат: создано новых={new_count_ref['count']}")
            
            save_processed_order(order_number)
            loaded_count += 1
            app.logger.info(f"Заказ {order_number} успешно обработан")
        
        if new_products_count:
            app.logger.info(f"Сохраняем {new_products_count} новых товаров")
            save_products(products)
        
        save_stocks(stocks)
        
        if composite_log:
            log_file = os.path.join(app.config['DATA_FOLDER'], 'ozon_composite_articles_log.json')
            try:
                with open(log_file, 'w', encoding='utf-8') as f:
                    json.dump(composite_log, f, indent=2, ensure_ascii=False)
                app.logger.info(f"Лог составных артикулов Ozon сохранен в {log_file}")
            except Exception as e:
                app.logger.error(f"Ошибка сохранения лога: {e}")
        
        app.logger.info("=" * 60)
        app.logger.info(f"ИТОГИ ЗАГРУЗКИ: загружено={loaded_count}, пропущено={skipped_count}, новых товаров={new_products_count}")
        app.logger.info("=" * 60)
        
        metrics.record_operation('ozon_orders_loaded')
        return jsonify({
            'success': True,
            'loaded': loaded_count,
            'skipped': skipped_count,
            'new_products': new_products_count,
            'composite_articles': composite_count,
            'composite_log': composite_log[:20],
            'missing_products': missing_products[:10]
        })
        
    except Exception as e:
        app.logger.error(f"Ошибка при загрузке заказов Ozon: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==================== ДИАГНОСТИЧЕСКИЙ МАРШРУТ ====================

@app.route('/debug/check-article', methods=['POST'])
@login_required
def debug_check_article():
    """Диагностическая функция для проверки поиска артикулов"""
    try:
        data = request.json
        composite_article = data.get('article', '')
        
        if not composite_article:
            return jsonify({'success': False, 'error': 'Артикул не указан'})
        
        products = load_products()
        
        existing_articles = [p['article'] for p in products]
        
        exact_match = None
        for p in products:
            if p['article'] == composite_article:
                exact_match = p
                break
        
        found_products, not_found = find_matching_products_enhanced(composite_article, products)
        
        specific_check = None
        if composite_article == '10ZPSR_PN_18_A_PEDZ_52_C' or '_' in composite_article:
            parts = composite_article.split('_')
            if len(parts) >= 2:
                part1 = parts[0]
                part2 = '_'.join(parts[1:])
                
                part1_exists = any(p['article'] == part1 for p in products)
                part2_exists = any(p['article'] == part2 for p in products)
                
                similar_to_part1 = []
                for p in products:
                    if p['article'].startswith(part1) and p['article'] != part1:
                        similar_to_part1.append(p['article'])
                
                specific_check = {
                    'part1': {
                        'article': part1,
                        'exists': part1_exists,
                        'similar': similar_to_part1[:5],
                        'product': next((p for p in products if p['article'] == part1), None)
                    },
                    'part2': {
                        'article': part2,
                        'exists': part2_exists,
                        'product': next((p for p in products if p['article'] == part2), None)
                    }
                }
        
        result = {
            'success': True,
            'article': composite_article,
            'exact_match': {
                'found': exact_match is not None,
                'product': exact_match
            },
            'existing_articles': existing_articles[:50],
            'search_result': {
                'found': [{'article': p['article'], 'name': p['name']} for p in found_products],
                'not_found': not_found
            },
            'specific_check': specific_check,
            'note': 'Часть 1 (10ZPSR) не найдена, но есть похожий артикул 10ZPSR_720'
        }
        
        return jsonify(result)
        
    except Exception as e:
        app.logger.error(f"Ошибка в диагностике: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ ОПЕРАЦИЙ ====================

@app.route('/operation/add', methods=['POST'])
@login_required
def add_operation():
    try:
        data = request.json
        operation_type = data.get('type')
        article = data.get('article')
        quantity = int(data.get('quantity', 0))
        order_number = data.get('order_number', 'Ручной ввод')
        date = data.get('date', datetime.now().strftime("%Y-%m-%d"))
        
        if not operation_type or not article or quantity <= 0:
            return jsonify({'success': False, 'error': 'Заполните все поля'})
        
        products = load_products()
        stocks = load_stocks()
        
        if '_' in article:
            app.logger.info(f"Обнаружен составной артикул при ручном добавлении: {article}")
            
            found_products, not_found_parts = find_matching_products_enhanced(article, products)
            
            if not found_products and not_found_parts:
                return jsonify({
                    'success': False, 
                    'error': f'Не найден ни один товар в составном артикуле. Ненайденные части: {", ".join(not_found_parts)}'
                })
            
            results = []
            new_products_count = 0
            
            for product in found_products:
                current_stock = stocks.get(product['article'], 0)
                if operation_type == "Расход" and current_stock < quantity:
                    results.append({
                        'article': product['article'],
                        'name': product['name'],
                        'status': 'error',
                        'message': f'Недостаточно товара. Доступно: {current_stock}'
                    })
                    continue
                
                order_data = {
                    "Operation Type": operation_type,
                    "Date": date,
                    "Order Number": order_number,
                    "Product Name": product['name'],
                    "Product Article": product['article'],
                    "Quantity": quantity
                }
                
                if save_order(order_data):
                    if operation_type == "Приход":
                        stocks[product['article']] = current_stock + quantity
                    elif operation_type == "Расход":
                        stocks[product['article']] = current_stock - quantity
                    
                    results.append({
                        'article': product['article'],
                        'name': product['name'],
                        'status': 'success',
                        'new_stock': stocks[product['article']]
                    })
            
            for part in not_found_parts:
                existing = False
                for p in products:
                    if p['article'] == part:
                        existing = True
                        break
                
                if not existing:
                    new_product = {
                        'name': f"Товар из комплекта {part}",
                        'article': part
                    }
                    products.append(new_product)
                    new_products_count += 1
                    
                    order_data = {
                        "Operation Type": operation_type,
                        "Date": date,
                        "Order Number": order_number,
                        "Product Name": new_product['name'],
                        "Product Article": part,
                        "Quantity": quantity
                    }
                    
                    if save_order(order_data):
                        if operation_type == "Приход":
                            stocks[part] = quantity
                        elif operation_type == "Расход":
                            stocks[part] = -quantity
                        
                        results.append({
                            'article': part,
                            'name': new_product['name'],
                            'status': 'success',
                            'new_stock': stocks[part],
                            'new_product': True
                        })
                else:
                    for p in products:
                        if p['article'] == part:
                            current_stock = stocks.get(part, 0)
                            order_data = {
                                "Operation Type": operation_type,
                                "Date": date,
                                "Order Number": order_number,
                                "Product Name": p['name'],
                                "Product Article": part,
                                "Quantity": quantity
                            }
                            if save_order(order_data):
                                if operation_type == "Приход":
                                    stocks[part] = current_stock + quantity
                                elif operation_type == "Расход":
                                    stocks[part] = current_stock - quantity
                                
                                results.append({
                                    'article': part,
                                    'name': p['name'],
                                    'status': 'success',
                                    'new_stock': stocks[part]
                                })
                            break
            
            if new_products_count > 0:
                save_products(products)
            
            save_stocks(stocks)
            
            success_count = sum(1 for r in results if r['status'] == 'success')
            error_count = len(results) - success_count
            
            metrics.record_operation('manual_operation_composite')
            return jsonify({
                'success': True,
                'message': f'Обработано составных артикулов: {len(results)} (успешно: {success_count}, ошибок: {error_count})',
                'results': results,
                'new_products': new_products_count
            })
        
        else:
            product = None
            for p in products:
                if p['article'] == article:
                    product = p
                    break
            
            if not product:
                return jsonify({'success': False, 'error': f'Товар с артикулом {article} не найден'})
            
            current_stock = stocks.get(article, 0)
            if operation_type == "Расход" and current_stock < quantity:
                return jsonify({'success': False, 'error': f'Недостаточно товара на складе. Доступно: {current_stock}'})
            
            order_data = {
                "Operation Type": operation_type,
                "Date": date,
                "Order Number": order_number,
                "Product Name": product['name'],
                "Product Article": article,
                "Quantity": quantity
            }
            
            if not save_order(order_data):
                return jsonify({'success': False, 'error': 'Ошибка при сохранении операции'})
            
            if operation_type == "Приход":
                stocks[article] = current_stock + quantity
            elif operation_type == "Расход":
                stocks[article] = current_stock - quantity
            elif operation_type == "Инвентаризация":
                stocks[article] = quantity
            
            save_stocks(stocks)
            
            metrics.record_operation('manual_operation')
            return jsonify({
                'success': True,
                'message': f'Операция "{operation_type}" успешно добавлена',
                'new_stock': stocks[article]
            })
        
    except Exception as e:
        app.logger.error(f"Ошибка при добавлении операции: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)})


# ==================== МАРШРУТЫ ЭКСПОРТА ====================

@app.route('/export/inventory')
@login_required
def export_inventory():
    try:
        products = load_products()
        stocks = load_stocks()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Инвентаризация"
        
        headers = ["№", "Артикул", "Название товара", "Текущий остаток", "Фактический остаток", "Расхождение", "Примечание"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        for row_num, product in enumerate(products, 2):
            article = product['article']
            name = product['name']
            current_stock = stocks.get(article, 0)
            
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=article)
            ws.cell(row=row_num, column=3, value=name)
            ws.cell(row=row_num, column=4, value=current_stock)
            ws.cell(row=row_num, column=6, value=f"=E{row_num}-D{row_num}")
        
        filename = f"инвентаризация_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/export/orders/<order_number>')
@login_required
def export_order(order_number):
    try:
        orders = load_orders()
        order_items = [o for o in orders if o.get('Order Number') == order_number]
        
        if not order_items:
            flash('Заказ не найден', 'danger')
            return redirect(url_for('orders_list'))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Заказ {order_number}"
        
        headers = ["№", "Артикул", "Название товара", "Количество"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        for row_num, item in enumerate(order_items, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=item.get('Product Article', ''))
            ws.cell(row=row_num, column=3, value=item.get('Product Name', ''))
            try:
                quantity = int(item.get('Quantity', 0))
            except:
                quantity = 0
            ws.cell(row=row_num, column=4, value=quantity)
        
        filename = f"order_{order_number}.xlsx"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash('Ошибка при экспорте заказа', 'danger')
        return redirect(url_for('orders_list'))


@app.route('/export/products/<format>')
@login_required
def export_products(format):
    """Экспорт товаров в различных форматах"""
    products = load_products()
    stocks = load_stocks()
    
    data = []
    for product in products:
        data.append({
            'name': product['name'],
            'article': product['article'],
            'model': product.get('model', ''),
            'stock': stocks.get(product['article'], 0)
        })
    
    if format == 'csv':
        return export_as_csv(data)
    elif format == 'excel':
        return export_as_excel(data)
    elif format == 'json':
        return export_as_json(data)
    else:
        return jsonify({'error': 'Unsupported format'}), 400


def export_as_csv(data):
    """Экспорт в CSV"""
    import csv
    from io import StringIO
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=['name', 'article', 'model', 'stock'])
    writer.writeheader()
    writer.writerows(data)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=products_{datetime.now().strftime("%Y%m%d")}.csv'
    return response


def export_as_excel(data):
    """Экспорт в Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"
    
    headers = ["Название", "Артикул", "Модель", "Остаток"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=item['name'])
        ws.cell(row=row_num, column=2, value=item['article'])
        ws.cell(row=row_num, column=3, value=item['model'])
        ws.cell(row=row_num, column=4, value=item['stock'])
    
    filename = f"products_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)


def export_as_json(data):
    """Экспорт в JSON"""
    return jsonify(data)


# ==================== API МАРШРУТЫ ====================

@app.route('/api/stats')
@login_required
def api_stats():
    try:
        products = load_products()
        stocks = load_stocks()
        orders = load_orders()
        
        return jsonify({
            'total_products': len(products),
            'total_orders': len(set([o.get('Order Number', '') for o in orders])),
            'total_stock': sum(stocks.values()),
            'low_stock': sum(1 for p in products if stocks.get(p['article'], 0) < 10)
        })
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/api/orders/calendar')
@login_required
def orders_calendar():
    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        
        start_date = datetime(year, month, 1).date()
        if month == 12:
            end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        
        orders = load_orders()
        
        daily_stats = {}
        for order in orders:
            try:
                order_date = datetime.strptime(order['Date'], '%Y-%m-%d').date()
                if start_date <= order_date <= end_date:
                    date_str = order_date.strftime('%Y-%m-%d')
                    if date_str not in daily_stats:
                        daily_stats[date_str] = {
                            'date': date_str,
                            'total_orders': 0,
                            'total_items': 0,
                            'orders': []
                        }
                    
                    order_num = order.get('Order Number', '')
                    if order_num not in daily_stats[date_str]['orders']:
                        daily_stats[date_str]['orders'].append(order_num)
                        daily_stats[date_str]['total_orders'] += 1
                    
                    daily_stats[date_str]['total_items'] += int(order.get('Quantity', 0))
                    
            except:
                continue
        
        return jsonify({
            'success': True,
            'data': list(daily_stats.values())
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/metrics')
@login_required
def show_metrics():
    """Показать метрики системы"""
    return jsonify(metrics.get_stats())


# ==================== МАРШРУТЫ БЭКАПА ====================

@app.route('/backup', methods=['POST'])
@login_required
def create_backup():
    success, result = backup_data()
    if success:
        metrics.record_operation('backup_created')
        return send_file(result, as_attachment=True, download_name=os.path.basename(result))
    else:
        return jsonify({'success': False, 'error': result})


@app.route('/upload/async', methods=['POST'])
@login_required
def upload_async():
    """Асинхронная загрузка файла"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Файл не выбран'})
    
    # Сохраняем файл
    filename = secure_filename(f"async_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    # Отправляем на асинхронную обработку
    task_id = file_processor.submit(filepath, lambda p: parse_pdf_orders(p))
    
    return jsonify({
        'success': True,
        'task_id': task_id,
        'message': 'Файл отправлен на обработку'
    })


@app.route('/upload/status/<task_id>')
@login_required
def upload_status(task_id):
    """Статус асинхронной обработки"""
    result = file_processor.get_result(task_id)
    if result:
        return jsonify(result)
    return jsonify({'status': 'processing'})


# ==================== ОБРАБОТЧИКИ ОШИБОК ====================

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f"Внутренняя ошибка сервера: {error}")
    return render_template('500.html'), 500


if __name__ == '__main__':
    print("="*60)
    print("Складской учет - сервер запускается")
    print("="*60)
    print("Адрес: http://127.0.0.1:8080")
    print("Логин: admin")
    print("Пароль: admin")
    print("="*60)
    print("Новые возможности:")
    print("✅ Кэширование данных для ускорения")
    print("✅ Индексация товаров для быстрого поиска")
    print("✅ Отчет по составным артикулам")
    print("✅ Поиск по части артикула")
    print("✅ Асинхронная обработка файлов")
    print("✅ Системные метрики")
    print("="*60)
    print("Для остановки нажмите Ctrl+C")
    print("="*60)
    
    app.run(debug=True, host='0.0.0.0', port=8080)
